#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
比對兩個 participants JSON 的 famid 重複狀況，結果匯出成 xlsx。
A = db.json   (正式環境 / 基準)
B = test.json (測試環境)

輸出：compare_famid.xlsx，含 5 個工作表
  1) Summary       彙總
  2) Duplicates    各檔重複的 famid
  3) Both_Diff     共同 famid 且兩邊資料有差異（逐欄）
  4) Only_in_A     只在 db.json 出現的整筆資料
  5) Only_in_B     只在 test.json 出現的整筆資料

需求套件：openpyxl  (pip install openpyxl)
檔案位置：db.json / test.json 與本腳本同資料夾。
"""

import json
import os
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FILE_A = os.path.join(BASE_DIR, "db.json")    # 正式環境（基準）
FILE_B = os.path.join(BASE_DIR, "test.json")  # 測試環境
OUT_XLSX = os.path.join(BASE_DIR, "compare_famid.xlsx")
FAMID_KEY = "famid"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="4472C4")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=12)
BASE_FONT = Font(name=FONT_NAME)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_records(path):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for key in ("participants", "data", "records", "items", "list"):
            if isinstance(data.get(key), list):
                return data[key]
        for v in data.values():
            if isinstance(v, list):
                return v
        return [data]
    raise ValueError(f"無法解析 JSON 結構：{path}")


def famid_of(rec):
    return rec.get(FAMID_KEY) if isinstance(rec, dict) else None


def find_dups(records):
    counts = Counter(famid_of(r) for r in records if famid_of(r) is not None)
    return {fid: n for fid, n in counts.items() if n > 1}


def dedupe(records):
    seen = {}
    for r in records:
        fid = famid_of(r)
        if fid is not None and fid not in seen:
            seen[fid] = r
    return seen


def cellval(v):
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    return v


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autofit(ws, max_width=60):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 0
        for cell in col:
            if cell.value is not None:
                longest = max(longest, max(len(s) for s in str(cell.value).split("\n")))
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), max_width)


def write_table(ws, headers, rows, start_row=1):
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h).font = BASE_FONT
    style_header(ws, len(headers), row=start_row)
    for i, rowvals in enumerate(rows, start=start_row + 1):
        for j, v in enumerate(rowvals, start=1):
            cell = ws.cell(row=i, column=j, value=cellval(v))
            cell.font = BASE_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


def main():
    for p in (FILE_A, FILE_B):
        if not os.path.exists(p):
            raise SystemExit(f"找不到檔案：{p}")

    recs_a = load_records(FILE_A)
    recs_b = load_records(FILE_B)

    dups_a = find_dups(recs_a)
    dups_b = find_dups(recs_b)

    map_a = dedupe(recs_a)
    map_b = dedupe(recs_b)
    set_a, set_b = set(map_a), set(map_b)
    both = set_a & set_b
    only_a = sorted(set_a - set_b, key=str)
    only_b = sorted(set_b - set_a, key=str)

    identical, differ_rows = [], []
    for fid in sorted(both, key=str):
        ra, rb = map_a[fid], map_b[fid]
        keys = sorted(set(ra) | set(rb), key=str)
        diffs = [(k, ra.get(k, "<無此欄位>"), rb.get(k, "<無此欄位>"))
                 for k in keys if ra.get(k, "<無此欄位>") != rb.get(k, "<無此欄位>")]
        if diffs:
            for k, va, vb in diffs:
                differ_rows.append((fid, k, cellval(va), cellval(vb)))
        else:
            identical.append(fid)
    differ_famids = len({r[0] for r in differ_rows})

    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "famid 比對彙總（A=db.json 正式 / B=test.json 測試）"
    ws["A1"].font = TITLE_FONT
    summary = [
        ("項目", "數值"),
        ("A db.json 總筆數", len(recs_a)),
        ("B test.json 總筆數", len(recs_b)),
        ("A 重複 famid 種類數", len(dups_a)),
        ("B 重複 famid 種類數", len(dups_b)),
        ("A 去重後 famid 數", len(set_a)),
        ("B 去重後 famid 數", len(set_b)),
        ("共同 famid 數", len(both)),
        ("　- 資料完全相同", len(identical)),
        ("　- 資料有差異", differ_famids),
        ("只在 A (db.json)", len(only_a)),
        ("只在 B (test.json)", len(only_b)),
    ]
    write_table(ws, summary[0], summary[1:], start_row=3)
    autofit(ws)

    ws = wb.create_sheet("Duplicates")
    drows = [("db.json", fid, n) for fid, n in sorted(dups_a.items(), key=lambda x: str(x[0]))]
    drows += [("test.json", fid, n) for fid, n in sorted(dups_b.items(), key=lambda x: str(x[0]))]
    if not drows:
        drows = [("（無）", "兩檔皆無重複 famid", "")]
    write_table(ws, ["檔案", "famid", "出現次數"], drows)
    autofit(ws)

    ws = wb.create_sheet("Both_Diff")
    if not differ_rows:
        differ_rows = [("（無）", "共同 famid 資料皆相同", "", "")]
    write_table(ws, ["famid", "欄位", "A 值 (db.json)", "B 值 (test.json)"], differ_rows)
    autofit(ws)

    def write_only(sheet_name, fids, src_map, src_label):
        ws = wb.create_sheet(sheet_name)
        if not fids:
            write_table(ws, ["famid", "資料"], [("（無）", f"沒有只在 {src_label} 出現的 famid")])
            autofit(ws)
            return
        cols = []
        for fid in fids:
            for k in src_map[fid]:
                if k not in cols:
                    cols.append(k)
        if FAMID_KEY in cols:
            cols.remove(FAMID_KEY)
        headers = [FAMID_KEY] + cols
        rows = [[fid] + [src_map[fid].get(k, "") for k in cols] for fid in fids]
        write_table(ws, headers, rows)
        autofit(ws)

    write_only("Only_in_A", only_a, map_a, "db.json")
    write_only("Only_in_B", only_b, map_b, "test.json")

    wb.save(OUT_XLSX)
    print(f"已輸出：{OUT_XLSX}")
    print(f"共同 {len(both)}（相同 {len(identical)} / 差異 {differ_famids}）"
          f"｜只在 A {len(only_a)}｜只在 B {len(only_b)}"
          f"｜重複 A {len(dups_a)} / B {len(dups_b)}")


if __name__ == "__main__":
    main()