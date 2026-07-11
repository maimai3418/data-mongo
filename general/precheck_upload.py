#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
上傳前重複檢查（dry-run / double-check）。

在「不實際寫入 MongoDB」的前提下，拿一份新的 Excel 資料跟資料庫現有資料比對，
確認：
  1) 哪些紀錄已經上傳過（依 unique key role + famid + record_date 判斷，
     等同 importer.upsert_many 會 skip 的那些）。
  2) 已存在的紀錄，其「計畫代碼」與這次要套用的計畫代碼是否相符
     —— 因為 upsert 用 $setOnInsert，已存在的紀錄不會被覆蓋，
     若計畫不同，重傳並不會更新計畫，容易誤判。
  3) 各 famid 在資料庫已有的 record_date，協助確認日期是否相符。

完全沿用 main.py 的流程（input_project_code → select_collections →
read_xlsx → split_by_collection），所以檢查的就是「真的會被上傳」的那批資料，
不更動任何現有模組。

輸出：precheck_report.xlsx
用法：
  python general/precheck_upload.py                # 預設讀 import_data.xlsx
  python general/precheck_upload.py 其他資料.xlsx   # 指定要 double-check 的檔案
"""

import os
import sys

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import SHARED_FIELDS
from src.reader import read_xlsx
from src.transformer import split_by_collection
from src.importer import get_db
from src.utils.select_collections import select_collections
from src.utils.input_project_code import input_project_code
from src.utils.wait_and_retry import wait_and_retry

load_dotenv()

DEFAULT_FILEPATH = "import_data.xlsx"
OUT_XLSX = "precheck_report.xlsx"
PROJECT_FIELD = "research_project_code"

# ---- 報表樣式（沿用 famid_compare.py 的風格）----
FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="4472C4")
WARN_FILL = PatternFill("solid", start_color="FFF2CC")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=12)
BASE_FONT = Font(name=FONT_NAME)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def norm(v):
    """統一把 key 欄位轉成可比對的字串。"""
    if v is None:
        return ""
    return str(v).strip()


def key_of(doc):
    return tuple(norm(doc.get(f)) for f in SHARED_FIELDS)


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autofit(ws, max_width=50):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 0
        for cell in col:
            if cell.value is not None:
                longest = max(longest, max(len(s) for s in str(cell.value).split("\n")))
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), max_width)


def write_table(ws, headers, rows, start_row=1, warn_col=None, warn_value=None):
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h).font = BASE_FONT
    style_header(ws, len(headers), row=start_row)
    for i, rowvals in enumerate(rows, start=start_row + 1):
        flag = warn_col is not None and rowvals[warn_col] == warn_value
        for j, v in enumerate(rowvals, start=1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = BASE_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if flag:
                cell.fill = WARN_FILL
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


def fetch_existing(db, collection, famids):
    """抓資料庫中這些 famid 的現有紀錄，建立 key -> 既有紀錄 的查找表。"""
    if not famids:
        return {}, {}
    projection = {f: 1 for f in SHARED_FIELDS}
    projection[PROJECT_FIELD] = 1
    projection["_id"] = 0
    cursor = db[collection].find({"famid": {"$in": list(famids)}}, projection)

    by_key = {}
    dates_by_famid = {}
    for d in cursor:
        by_key[key_of(d)] = d
        fam = norm(d.get("famid"))
        dates_by_famid.setdefault(fam, set()).add(norm(d.get("record_date")))
    return by_key, dates_by_famid


def main():
    filepath = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_FILEPATH
    if not os.path.exists(filepath):
        raise SystemExit(f"找不到檔案：{filepath}")

    # 與 main.py 相同：要套用的計畫代碼（選填）
    project_code = input_project_code()
    new_project = norm(project_code)

    selected = select_collections()
    if selected is None:
        return

    print(f"reading file... ({filepath})")
    df = read_xlsx(filepath)
    print(f"total rows: {len(df)}")

    collections, error_rows, skipped_rows = split_by_collection(df, selected)

    db = get_db()

    summary_rows = []          # (collection, 檢查筆數, 新資料, 已存在, 計畫不符)
    dup_rows = []              # 已存在的紀錄
    new_rows = []              # 會被新增的紀錄
    mismatch_rows = []         # 已存在且計畫不符
    datecheck_rows = []        # famid 層級的日期比對

    for collection, docs in collections.items():
        if not docs:
            summary_rows.append((collection, 0, 0, 0, 0))
            continue

        famids = {norm(d.get("famid")) for d in docs}
        by_key, dates_by_famid = fetch_existing(db, collection, famids)

        n_new = n_dup = n_mismatch = 0
        seen_famids = set()

        for doc in docs:
            role, famid, rdate = norm(doc.get("role")), norm(doc.get("famid")), norm(doc.get("record_date"))
            existing = by_key.get((role, famid, rdate))

            if existing is None:
                n_new += 1
                new_rows.append((collection, role, famid, rdate, new_project or "（未填）"))
            else:
                n_dup += 1
                db_project = norm(existing.get(PROJECT_FIELD))
                # 只有當這次有指定計畫代碼時才比對計畫是否相符
                if new_project and new_project != db_project:
                    match = "不符"
                    n_mismatch += 1
                    mismatch_rows.append(
                        (collection, role, famid, rdate, db_project or "（DB 未記錄）", new_project)
                    )
                elif not new_project:
                    match = "未指定計畫"
                else:
                    match = "相符"
                dup_rows.append(
                    (collection, role, famid, rdate, db_project or "（DB 未記錄）", new_project or "（未填）", match)
                )

            # famid 層級日期比對（每個 famid 只記一次）
            fam_token = (collection, famid)
            if fam_token not in seen_famids:
                seen_famids.add(fam_token)
                new_dates = sorted({norm(d.get("record_date")) for d in docs if norm(d.get("famid")) == famid})
                db_dates = sorted(dates_by_famid.get(famid, set()))
                only_new = [d for d in new_dates if d not in db_dates]
                datecheck_rows.append((
                    collection, famid,
                    ", ".join(new_dates),
                    ", ".join(db_dates) if db_dates else "（DB 無此 famid）",
                    ", ".join(only_new) if only_new else "（無新日期）",
                ))

        summary_rows.append((collection, len(docs), n_new, n_dup, n_mismatch))

    # ---- 輸出 Excel ----
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"上傳前重複檢查彙總（檔案：{os.path.basename(filepath)}｜計畫：{new_project or '未指定'}）"
    ws["A1"].font = TITLE_FONT
    tot = [sum(r[i] for r in summary_rows) for i in range(1, 5)]
    write_table(
        ws,
        ["量表", "檢查筆數", "新資料(會上傳)", "已存在(會略過)", "計畫不符"],
        summary_rows + [("總計", *tot)],
        start_row=3,
    )
    autofit(ws)

    ws = wb.create_sheet("Duplicates")
    if not dup_rows:
        dup_rows = [("（無）", "", "", "", "", "", "沒有重複的紀錄")]
    write_table(
        ws,
        ["量表", "role", "famid", "record_date", "DB 計畫代碼", "本次計畫代碼", "計畫是否相符"],
        dup_rows, warn_col=6, warn_value="不符",
    )
    autofit(ws)

    ws = wb.create_sheet("Project_Mismatch")
    if not mismatch_rows:
        mismatch_rows = [("（無）", "", "", "", "", "計畫全部相符")]
    write_table(
        ws,
        ["量表", "role", "famid", "record_date", "DB 既有計畫", "本次要套用計畫"],
        mismatch_rows,
    )
    autofit(ws)

    ws = wb.create_sheet("New_Rows")
    if not new_rows:
        new_rows = [("（無）", "", "", "", "沒有新資料")]
    write_table(ws, ["量表", "role", "famid", "record_date", "本次計畫代碼"], new_rows)
    autofit(ws)

    ws = wb.create_sheet("Date_Check")
    if not datecheck_rows:
        datecheck_rows = [("（無）", "", "", "", "")]
    write_table(
        ws,
        ["量表", "famid", "本檔 record_date", "DB 既有 record_date", "本檔多出的日期"],
        datecheck_rows,
    )
    autofit(ws)

    wait_and_retry(lambda: wb.save(OUT_XLSX), OUT_XLSX)

    # ---- 主控台摘要 ----
    print("\n===== 檢查結果 =====")
    for collection, total, n_new, n_dup, n_mismatch in summary_rows:
        if total == 0:
            continue
        warn = f"｜⚠️ 計畫不符 {n_mismatch}" if n_mismatch else ""
        print(f"[{collection}] 檢查 {total}｜新資料 {n_new}｜已存在 {n_dup}{warn}")
    print(f"\n總計：檢查 {tot[0]}｜新資料 {tot[1]}｜已存在 {tot[2]}｜計畫不符 {tot[3]}")
    print(f"已輸出報表：{OUT_XLSX}")


if __name__ == "__main__":
    main()
