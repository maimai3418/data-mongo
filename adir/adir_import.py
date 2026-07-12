#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ADI-R 匯入：讀取 ADI-R xlsx，匯入單一 ADIR collection。

與 ADOS 匯入（ados_import.py）的差異：
  - 單一 collection（ADIR），不分 module。
  - null 轉換：item 欄位 999→null（同 ADOS），但月齡欄位（field.json
    range 涵蓋 999，如 0–999）例外，其 999 為合法值、先照 range 驗證。
    空白儲存格→null 的邏輯不變。
  - version 欄位（full/current）於匯入前的 rename 階段（adir_field_rename.py）
    依工作表名稱標註；此處僅視為一般 header 欄位帶入。

共通邏輯：
  - 相同 famid + record_date 視為重複，不重複匯入（$setOnInsert）。
  - item 欄位全為空的列略過不上傳。
  - 錯誤匯出至 {today}_adir_import_error.xlsx。

用法（從專案根目錄執行）：
  python adir/adir_import.py <檔案或資料夾>
  python adir/adir_import.py data_dir/
  python adir/adir_import.py a.xlsx b.xlsx
  python adir/adir_import.py data_dir/ --sheets Full Current
  python adir/adir_import.py data_dir/ --dry-run
  python adir/adir_import.py a.xlsx --project-code NHRI113
  python adir/adir_import.py a.xlsx -e errors.xlsx

需求套件：pandas、openpyxl、pymongo、python-dotenv
"""

import os
import sys
import argparse
import warnings
from datetime import date

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

import pandas as pd
from dotenv import load_dotenv
from pymongo import UpdateOne
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.importer import get_db
from src.utils.wait_and_retry import wait_and_retry
from adir_config import (
    ADIR_COLLECTION,
    ADIR_SHARED_FIELDS,
    ADIR_HEADER_FIELDS,
    ADIR_NULL_VALUE,
    ADIR_FIELDS_DIR,
    load_field_rules,
)

try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

load_dotenv()

TODAY = date.today().strftime("%Y%m%d")
DEFAULT_ERROR_OUTPUT = f"{TODAY}_adir_import_error.xlsx"
PROJECT_FIELD = "research_project_code"


# ── 數值處理 ────────────────────────────────────────────────────────────
def is_blank(val):
    return val is None or (isinstance(val, float) and pd.isna(val))


def to_number(s):
    try:
        f = float(s)
    except (ValueError, TypeError):
        return None
    return int(f) if f.is_integer() else f


def clean_value(val):
    if is_blank(val):
        return None
    s = str(val).strip()
    return s if s != "" else None


def to_date_text(col, val):
    if is_blank(val):
        return None, None
    s = str(val).strip()
    if s == "":
        return None, None
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        dt = pd.to_datetime(s, errors="coerce")
    if pd.isna(dt):
        return s, f"{col} 日期格式錯誤"
    return dt.strftime("%Y-%m-%d"), None


def allows_null_value(rule):
    """判斷欄位的合法 range 是否涵蓋 null sentinel（999）。

    月齡欄位（range 如 0–999）的 999 為合法值，不應被當成遺漏值轉 null。
    """
    if not rule:
        return False
    rng = rule.get("range")
    return bool(rng) and rng[0] <= ADIR_NULL_VALUE <= rng[1]


def convert_item(col, val, rule):
    """item 欄位轉換，回傳 (converted, error)。

    999 → null（同 ADOS），但月齡欄位（range 涵蓋 999）例外，其 999 為合法值、
    照 range 驗證。空值→null；有 rule 時做型別 + 範圍驗證；無 rule 時原樣上傳。
    """
    if is_blank(val):
        return None, None
    s = str(val).strip()
    if s == "":
        return None, None

    num = to_number(s)

    # 999 → null，除非該欄位 range 合法涵蓋 999（月齡欄位）
    if num is not None and num == ADIR_NULL_VALUE and not allows_null_value(rule):
        return None, None

    # 無規則：原始值上傳（能轉數字就存數字）
    if rule is None:
        return (num if num is not None else s), None

    rtype = rule.get("type")
    if rtype in ("int", "float"):
        if num is None:
            return None, f"{col} 數值錯誤"
        if rtype == "int" and not isinstance(num, int):
            return None, f"{col} 數值錯誤(非整數)"
        if "range" in rule:
            lo, hi = rule["range"]
            if not (lo <= num <= hi):
                return None, f"{col} 數值錯誤({num} 不在 {lo}–{hi})"
        return num, None

    return (num if num is not None else s), None


# ── 單一工作表處理 ──────────────────────────────────────────────────────
def process_sheet(filename, sheet_name, df, errors, field_rules):
    """處理單一工作表，回傳 (docs, n_empty)。"""
    df = df.copy()
    df.columns = df.columns.str.strip()
    df = df.loc[:, df.columns.notna()]
    df = df.dropna(how="all")

    header_set = set(ADIR_HEADER_FIELDS)
    item_cols = [c for c in df.columns if c not in header_set]
    header_cols = [c for c in df.columns if c in header_set]

    docs = []
    n_empty = 0

    for _, row in df.iterrows():
        rowd = row.to_dict()
        doc = {}
        row_errors = []

        # ── header 欄位 ──
        for col in header_cols:
            rule = field_rules.get(col)
            if rule and rule.get("type") == "date":
                new_val, err = to_date_text(col, rowd.get(col))
                if err:
                    row_errors.append(err)
                doc[col] = new_val
            else:
                doc[col] = clean_value(rowd.get(col))

        # ── item 欄位 ──
        for col in item_cols:
            rule = field_rules.get(col)
            if rule and rule.get("type") == "date":
                new_val, err = to_date_text(col, rowd.get(col))
                if err:
                    row_errors.append(err)
                doc[col] = new_val
            else:
                new_val, err = convert_item(col, rowd.get(col), rule)
                if err:
                    row_errors.append(err)
                else:
                    doc[col] = new_val

        # item 欄位全為空 → 略過
        if item_cols and all(doc.get(c) is None for c in item_cols):
            n_empty += 1
            continue

        # 必要欄位檢查
        missing = [
            f for f in ADIR_SHARED_FIELDS
            if doc.get(f) is None or str(doc.get(f)).strip() == ""
        ]
        if missing:
            row_errors.append(f"缺少必要欄位: {', '.join(missing)}")

        if row_errors:
            errors.append({
                "file": filename,
                "sheet": sheet_name,
                "famid": clean_value(rowd.get("famid")),
                "record_date": clean_value(rowd.get("record_date")),
                "error": "; ".join(row_errors),
            })
        else:
            docs.append(doc)

    return docs, n_empty


# ── 寫入 DB ─────────────────────────────────────────────────────────────
def upsert_adir(db, docs):
    """以 famid + record_date 為 unique key 上傳；重複者略過。"""
    if not docs:
        return 0, 0
    operations = [
        UpdateOne(
            {k: doc.get(k) for k in ADIR_SHARED_FIELDS},
            {"$setOnInsert": doc},
            upsert=True,
        )
        for doc in docs
    ]
    result = db[ADIR_COLLECTION].bulk_write(operations)
    inserted = result.upserted_count
    skipped = len(docs) - inserted
    return inserted, skipped


# ── 錯誤匯出 ────────────────────────────────────────────────────────────
def export_errors(errors, output_path):
    if not errors:
        print("\n✓ 無錯誤，不產生 error 檔案")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "errors"

    headers = ["file", "sheet", "famid", "record_date", "error"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="C0392B")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, err in enumerate(errors, 2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=err.get(key))

    widths = {"A": 36, "B": 14, "C": 14, "D": 14, "E": 60}
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"

    wait_and_retry(lambda: wb.save(output_path), output_path)
    print(f"\n✗ 共 {len(errors)} 筆錯誤，已匯出: {os.path.basename(output_path)}")


# ── 檔案蒐集 ────────────────────────────────────────────────────────────
def collect_files(paths):
    files = []
    for p in paths:
        if os.path.isdir(p):
            for f in sorted(os.listdir(p)):
                if (f.endswith(".xlsx") and not f.startswith("~$")
                        and not f.startswith(TODAY)):
                    files.append(os.path.join(p, f))
        elif os.path.isfile(p):
            files.append(p)
        else:
            print(f"⚠️  略過：找不到 {p}")
    seen, unique = set(), []
    for f in files:
        ap = os.path.abspath(f)
        if ap not in seen:
            seen.add(ap)
            unique.append(f)
    return unique


def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="ADI-R 匯入：讀取 xlsx 匯入 ADIR collection")
    parser.add_argument("paths", nargs="+",
                        help="要匯入的 xlsx 檔案或資料夾")
    parser.add_argument("--sheets", nargs="+", default=None, metavar="S",
                        help="只匯入指定工作表（預設全部）")
    parser.add_argument("--project-code", default=None, metavar="CODE",
                        help="替每筆資料標註計畫代碼（寫入 research_project_code）")
    parser.add_argument("--dry-run", action="store_true",
                        help="只驗證與轉換、不寫入 MongoDB")
    parser.add_argument("-e", "--error-output", default=DEFAULT_ERROR_OUTPUT,
                        metavar="XLSX",
                        help=f"錯誤報告輸出路徑（預設 {DEFAULT_ERROR_OUTPUT}）")
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)

    files = collect_files(args.paths)
    if not files:
        print("Error: 找不到任何 xlsx 檔案")
        sys.exit(1)

    sheet_filter = set(args.sheets) if args.sheets else None
    project_code = (args.project_code or "").strip() or None

    field_rules = load_field_rules()

    n_ranged = sum(1 for r in field_rules.values() if r and "range" in r)
    print(f"📄 檔案數：{len(files)}｜collection：{ADIR_COLLECTION}"
          f"{'｜DRY-RUN（不寫入）' if args.dry_run else ''}")
    print(f"📐 欄位限制來源：{ADIR_FIELDS_DIR}")
    print(f"   有範圍限制的欄位：{n_ranged}")
    if sheet_filter:
        print(f"📑 只匯入工作表：{', '.join(sorted(sheet_filter))}")
    if project_code:
        print(f"🏷️  計畫代碼：{project_code}")

    errors = []
    all_docs = []
    total_empty = 0

    for filepath in files:
        filename = os.path.basename(filepath)
        print(f"\n處理: {filename}")
        try:
            all_sheets = pd.read_excel(filepath, sheet_name=None, dtype=str)
        except Exception as e:
            errors.append({"file": filename, "sheet": "-", "famid": "",
                           "record_date": "", "error": f"無法讀取檔案: {e}"})
            continue

        for sheet_name, df in all_sheets.items():
            if sheet_filter and sheet_name not in sheet_filter:
                print(f"  [{sheet_name}] 未在 --sheets 指定，略過")
                continue

            docs, n_empty = process_sheet(
                filename, sheet_name, df, errors, field_rules)

            if project_code:
                for doc in docs:
                    doc[PROJECT_FIELD] = project_code

            all_docs.extend(docs)
            total_empty += n_empty
            print(f"  [{sheet_name}] 有效 {len(docs)}｜略過(空值) {n_empty}")

    # ── 寫入 MongoDB ──
    print("\n===== 匯入結果 =====")
    n_err = len(errors)
    if args.dry_run:
        print(f"[{ADIR_COLLECTION}] dry-run：待上傳 {len(all_docs)}｜"
              f"略過(空值) {total_empty}｜錯誤 {n_err}")
    else:
        db = get_db()
        inserted, skipped = upsert_adir(db, all_docs)
        print(f"[{ADIR_COLLECTION}] 新增 {inserted}｜重複略過 {skipped}｜"
              f"略過(空值) {total_empty}｜錯誤 {n_err}")

    export_errors(errors, args.error_output)
    print("\ndone")


if __name__ == "__main__":
    main()