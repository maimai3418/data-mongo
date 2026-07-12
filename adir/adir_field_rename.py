#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ADI-R 欄位驗證 & 重命名腳本（匯入前的前置處理）

與 ADOS（ados_field_rename.py）的差異：
  - 單一 field.json（扁平結構），不分 module。
  - 處理檔案內「所有工作表」（Full / Current 等），非固定 M1~M4。
  - item 欄位名稱本身即為 field.json 的標準名（adir002、adire11…），
    只做大小寫正規化，不做前綴改名。

處理內容（對每個工作表）：
  1. famid：缺少時嘗試從 family+id 或 site+fam+id 組合。
  2. record_date：缺少時嘗試從 int_y/int_yr + int_m + int_d 組合。
  3. header 欄位改名：
       p_intloc → interview_location
       p_interv → interviewer
     其餘欄位（含 item）做大小寫正規化至 field.json 標準名。
  4. version：依工作表名稱（ADIR_SHEET_VERSION_MAP）標註 full / current；
     資料本身已有 version 者不覆蓋（僅補空白）。
  5. record_date / birth_date 統一為 YYYY-MM-DD 文字格式。
  6. 只保留 field.json 定義的欄位，並依 field.json 順序排列
     （family、id、int_y 等組合來源欄位在轉換後會被移除）。
  7. 錯誤匯出至 {today}_adir_rename_error.xlsx。

用法（從專案根目錄執行）：
  # 預設讀 adir/fields 下的 ADIR field.json
  python adir/adir_field_rename.py <檔案或資料夾>
  python adir/adir_field_rename.py data_dir/
  python adir/adir_field_rename.py a.xlsx b.xlsx

  # 指定其他 field.json
  python adir/adir_field_rename.py data_dir/ --field-json path/to/ADIR_fields.json

  # 不覆蓋原檔，另存 {stem}_renamed.xlsx
  python adir/adir_field_rename.py data_dir/ --no-overwrite

  # 其他選項
  python adir/adir_field_rename.py data_dir/ --famid-prefix 4
  python adir/adir_field_rename.py a.xlsx -e errors.xlsx

需求套件：pandas、numpy、openpyxl
"""

import os
import sys
import json
import argparse
import pandas as pd
import numpy as np
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 讓本檔可 import 同層的 adir_config（沿用其 field.json 偵測與 header 定義），
# 以及專案根目錄的 src 套件（wait_and_retry 等）
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_SCRIPT_DIR)
if _SCRIPT_DIR not in sys.path:
    sys.path.insert(0, _SCRIPT_DIR)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from adir_config import ADIR_FIELDS_DIR, discover_field_json, ADIR_SHEET_VERSION_MAP
from src.utils.wait_and_retry import wait_and_retry

# 讓中文/emoji 在 Windows 主控台（cp950）也能正常輸出
try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

# ═══════════════════════════════════════════════════════════════
# CONFIG（預設值；皆可由命令列覆寫）
# ═══════════════════════════════════════════════════════════════
TODAY = date.today().strftime("%Y%m%d")

# famid 組合公式中的前綴常數
# famid = FAMID_PREFIX * 100000 + site * 10000 + fam * 10 + id
DEFAULT_FAMID_PREFIX = 4

# 欄位名稱 variants（case-insensitive 比對）
SITE_COL = "site"
FAM_COL = "fam"
ID_COL = "id"
FAMILY_COL = "family"

INT_Y_VARIANTS = ["int_yr", "int_y"]
INT_M_COL = "int_m"
INT_D_COL = "int_d"

# header 舊欄名 → 新欄名（key 為小寫，比對 case-insensitive）
HEADER_RENAMES = {
    "p_intloc": "interview_location",
    "p_interv": "interviewer",
}

# 日期欄位（改名後的新名稱），統一為 YYYY-MM-DD 文字
DATE_FIELDS = ["record_date", "birth_date"]
# ═══════════════════════════════════════════════════════════════


def load_field_list(field_json_path):
    """讀取 ADIR field.json，回傳欄位名稱清單（保留原順序）。"""
    with open(field_json_path, "r", encoding="utf-8") as f:
        fields = json.load(f)
    return list(fields.keys())


def build_famid_series(df, famid_prefix):
    """嘗試從 family+id 或 site+fam+id 組合 famid，回傳 (Series|None, method_str)。"""
    cols_lower = {c.lower(): c for c in df.columns}

    # 方式 1: family + id
    fam_col = cols_lower.get(FAMILY_COL)
    id_col = cols_lower.get(ID_COL)
    if fam_col and id_col:
        family = pd.to_numeric(df[fam_col], errors="coerce")
        pid = pd.to_numeric(df[id_col], errors="coerce")
        if family.notna().any() and pid.notna().any():
            famid = (family * 10 + pid).astype("Int64").astype(str)
            famid = famid.str.replace("<NA>", "")
            return famid, "family*10+id"

    # 方式 2: site + fam + id
    site_col = cols_lower.get(SITE_COL)
    fam_col2 = cols_lower.get(FAM_COL)
    if site_col and fam_col2 and id_col:
        site = pd.to_numeric(df[cols_lower[SITE_COL]], errors="coerce")
        fam = pd.to_numeric(df[cols_lower[FAM_COL]], errors="coerce")
        pid = pd.to_numeric(df[cols_lower[ID_COL]], errors="coerce")
        if site.notna().any() and fam.notna().any() and pid.notna().any():
            famid = (
                famid_prefix * 100000 + site * 10000 + fam * 10 + pid
            ).astype("Int64").astype(str)
            famid = famid.str.replace("<NA>", "")
            return famid, f"{famid_prefix}*100000+site*10000+fam*10+id"

    return None, None


def normalize_date_col(series):
    """將日期欄位統一轉為 YYYY-MM-DD 文字格式，無法解析者保留原值。"""
    parsed = pd.to_datetime(series, errors="coerce", format="mixed")
    formatted = parsed.dt.strftime("%Y-%m-%d")
    # 無法解析的保留原始字串（或 NaN）
    return formatted.where(parsed.notna(), series)


def build_record_date_series(df):
    """嘗試從 int_y/int_yr + int_m + int_d 組合 record_date，回傳 YYYY-MM-DD 文字。"""
    cols_lower = {c.lower(): c for c in df.columns}

    y_col = None
    for v in INT_Y_VARIANTS:
        if v in cols_lower:
            y_col = cols_lower[v]
            break
    m_col = cols_lower.get(INT_M_COL)
    d_col = cols_lower.get(INT_D_COL)

    if y_col and m_col and d_col:
        y = pd.to_numeric(df[y_col], errors="coerce")
        m = pd.to_numeric(df[m_col], errors="coerce")
        d = pd.to_numeric(df[d_col], errors="coerce")
        dates = pd.to_datetime(
            {"year": y, "month": m, "day": d}, errors="coerce"
        )
        if dates.notna().any():
            text = dates.dt.strftime("%Y-%m-%d").where(dates.notna(), np.nan)
            return text, f"{y_col}+{m_col.lower()}+{d_col.lower()}"

    return None, None


def process_sheet(sheet_name, df, all_fields, famid_prefix):
    """處理單一工作表，回傳 (processed_df | None, errors, infos)。

    有 blocking error（famid / record_date 無法取得）時回傳 (None, errors, infos)。
    """
    df = df.copy()
    df.columns = df.columns.str.strip()
    df = df.loc[:, df.columns.notna()]

    cols_lower = {c.lower(): c for c in df.columns}
    errors = []
    infos = []

    # ── 1. famid ──
    famid_series = None
    famid_method = None
    if "famid" not in cols_lower:
        famid_series, famid_method = build_famid_series(df, famid_prefix)
        if famid_series is None:
            errors.append("缺少 famid，且無法從 family+id 或 site+fam+id 組合")

    # ── 2. record_date ──
    rdate_series = None
    rdate_method = None
    if "record_date" not in cols_lower:
        rdate_series, rdate_method = build_record_date_series(df)
        if rdate_series is None:
            errors.append("缺少 record_date，且無法從 int_y(r)/int_m/int_d 組合")

    if errors:
        return None, errors, infos

    # ── 3. 組合 famid / record_date（如需要）──
    if famid_series is not None:
        df.insert(0, "famid", famid_series)
        infos.append(f"famid 已從 {famid_method} 組合")

    if rdate_series is not None:
        insert_pos = list(df.columns).index("famid") + 1 if "famid" in df.columns else 1
        df.insert(insert_pos, "record_date", rdate_series)
        infos.append(f"record_date 已從 {rdate_method} 組合")

    # ── 4. 建立 rename map 並執行 ──
    cols_lower_current = {c.lower(): c for c in df.columns}
    rename_map = {}

    # header 舊→新（p_intloc / p_interv）
    for old_lower, new_name in HEADER_RENAMES.items():
        orig = cols_lower_current.get(old_lower)
        if orig:
            rename_map[orig] = new_name

    # 其餘欄位（含 item）大小寫正規化至 field.json 標準名
    for target in all_fields:
        orig = cols_lower_current.get(target.lower())
        if orig and orig != target and orig not in rename_map:
            rename_map[orig] = target

    if rename_map:
        df.rename(columns=rename_map, inplace=True)

    # ── 5. version：資料本身 > 工作表名稱推斷（依工作表名標註 full / current）──
    auto_version = ADIR_SHEET_VERSION_MAP.get(sheet_name.strip().lower())
    if auto_version:
        if "version" in df.columns:
            blank = df["version"].isna() | (df["version"].astype(str).str.strip() == "")
            if blank.any():
                df.loc[blank, "version"] = auto_version
                infos.append(
                    f"version 空白處補為 {auto_version}（依工作表名 {sheet_name}）")
        else:
            df["version"] = auto_version
            infos.append(f"version 已標註為 {auto_version}（依工作表名 {sheet_name}）")

    # ── 6. 日期欄位統一為 YYYY-MM-DD 文字格式 ──
    for date_col in DATE_FIELDS:
        if date_col in df.columns:
            df[date_col] = normalize_date_col(df[date_col])

    # ── 7. 只保留 field.json 定義的欄位，依 field.json 順序 ──
    keep = set(df.columns) & set(all_fields)
    ordered = [f for f in all_fields if f in keep]
    df = df[ordered]

    return df, errors, infos


def process_file(filepath, all_fields, errors, overwrite, famid_prefix):
    """處理單一 xlsx 檔案（所有工作表），回傳是否有任何 sheet 被成功處理。"""
    filename = os.path.basename(filepath)
    try:
        all_sheets = pd.read_excel(filepath, sheet_name=None, dtype=str)
    except Exception as e:
        errors.append({"file": filename, "sheet": "-", "error": f"無法讀取檔案: {e}"})
        return False

    modified = {}
    any_success = False

    for sheet_name, df in all_sheets.items():
        processed, sheet_errors, infos = process_sheet(
            sheet_name, df, all_fields, famid_prefix)

        for msg in infos:
            print(f"  [{sheet_name}] {msg}")

        if sheet_errors:
            for e in sheet_errors:
                errors.append({"file": filename, "sheet": sheet_name, "error": e})
            print(f"  [{sheet_name}] 略過（{len(sheet_errors)} 個錯誤）")
            continue

        modified[sheet_name] = processed
        any_success = True
        print(f"  [{sheet_name}] 已處理（{len(processed.columns)} 欄）")

    # ── 寫回檔案 ──
    if any_success:
        if overwrite:
            out_path = filepath
        else:
            stem = Path(filepath).stem
            out_path = os.path.join(os.path.dirname(filepath), f"{stem}_renamed.xlsx")

        def _save():
            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                for sheet_name, df in modified.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

        wait_and_retry(_save, out_path)
        print(f"  已儲存: {os.path.basename(out_path)} ({', '.join(modified.keys())})")

    return any_success


def export_errors(errors, error_output):
    """匯出錯誤報告至 Excel。"""
    if not errors:
        print("\n✓ 無錯誤，不產生 error 檔案")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "errors"

    headers = ["file", "sheet", "error"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="C0392B")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, err in enumerate(errors, 2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=err.get(key))

    widths = {"A": 40, "B": 16, "C": 60}
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"

    wait_and_retry(lambda: wb.save(error_output), error_output)
    print(f"\n✗ 共 {len(errors)} 筆錯誤，已匯出: {os.path.basename(error_output)}")


def collect_files(paths):
    """把檔案/資料夾參數展開成去重後的 xlsx 清單。"""
    files = []
    for p in paths:
        if os.path.isdir(p):
            for f in sorted(os.listdir(p)):
                if (f.endswith(".xlsx") and not f.startswith("~$")
                        and not f.startswith(TODAY)):
                    files.append(os.path.join(p, f))
        elif os.path.isfile(p):
            files.append(p)
        else:
            print(f"⚠️  略過：找不到 {p}")
    seen, unique = set(), []
    for f in files:
        ap = os.path.abspath(f)
        if ap not in seen:
            seen.add(ap)
            unique.append(f)
    return unique


def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="驗證並重命名 ADI-R xlsx（所有工作表）欄位為 field.json 標準格式")
    parser.add_argument("paths", nargs="+",
                        help="待處理 xlsx 檔案或資料夾")
    parser.add_argument("--field-json", default=None, metavar="FILE",
                        help="ADIR field.json 路徑（預設自動偵測 adir/fields）")
    parser.add_argument("--famid-prefix", type=int, default=DEFAULT_FAMID_PREFIX,
                        metavar="N",
                        help=f"famid 組合前綴常數（預設 {DEFAULT_FAMID_PREFIX}）")
    parser.add_argument("--no-overwrite", action="store_true",
                        help="不覆蓋原檔，改另存 {stem}_renamed.xlsx（預設覆蓋原檔）")
    parser.add_argument("-e", "--error-output", default=None, metavar="XLSX",
                        help="錯誤報告輸出路徑（預設 {today}_adir_rename_error.xlsx）")
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)

    field_json = args.field_json or discover_field_json()
    if not field_json or not os.path.isfile(field_json):
        print(f"Error: 找不到 ADIR field.json（搜尋於 {ADIR_FIELDS_DIR}），"
              f"或用 --field-json 指定。")
        sys.exit(1)

    files = collect_files(args.paths)
    if not files:
        print("Error: 找不到任何 xlsx 檔案")
        sys.exit(1)

    overwrite = not args.no_overwrite
    error_output = args.error_output or f"{TODAY}_adir_rename_error.xlsx"

    all_fields = load_field_list(field_json)
    print(f"使用 field.json：{os.path.basename(field_json)}（{len(all_fields)} 欄）")
    print(f"找到 {len(files)} 個檔案\n")

    errors = []
    for filepath in files:
        print(f"處理: {os.path.basename(filepath)}")
        process_file(filepath, all_fields, errors, overwrite, args.famid_prefix)

    export_errors(errors, error_output)
    print("\ndone")


if __name__ == "__main__":
    main()
