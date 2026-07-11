#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CANTAB 原始資料欄位改名（rename）。

依 cantab/fields/cantab_new_fields.xlsx Sheet1 的對照表（raw_col → db_code），
把原始資料檔的長欄名（如 "MOT Mean latency"）改成 DB 代碼（如 "MOTmL"），
輸出改名後的 xlsx 供後續 cantab_precheck.py / cantab_import.py 使用。

規則（欄名先 strip() 轉小寫比對，依序判斷）：
  1. 已是 db_code（field.json 欄位或對照表 db_code）→ 直接使用；
     大小寫/空白與標準寫法不同時修正成標準寫法並列入報告。
  2. 是對照表 raw_col 欄名 → 轉成 db_code（輸出用 db_code 原字串，
     大小寫與 field.json 完全一致）。header 欄位不在 Sheet1 上，
     依 cantab_config.HEADER_RENAMES 改名：Subject ID→famid、Gender→sex、
     Session start time→session_start_time、Age→raw_cantab_age、NART→nart。
  3. 對照表上有 raw_col 但 db_code 空白 = 明定不收 → 丟棄（列入報告；
     目前版本的對照表沒有這類列）。
  4. 都認不得的欄位 → 保留原名不改，報告標黃供人工比對
     （核心原則：非預期值必須明確浮現、模糊記錄保留絕不刪除）。
  - 只改欄名，不動任何儲存格值（famid 的 .0 殘留由 cantab_import.py 處理）。
  - 原始檔不動，資料不遺失。

輸出單位：
  - 傳入資料夾 → 該資料夾下所有檔案（含各工作表）改名後彙整成
    一個 xlsx（檔名 = 資料夾名稱），欄位取聯集、缺欄留空。
  - 傳入單一檔案 → 一檔一輸出（多工作表也彙整成單一工作表）。

用法（從專案根目錄執行）：
  python cantab/cantab_rename.py <檔案或資料夾>...
  python cantab/cantab_rename.py raw_dir/
  python cantab/cantab_rename.py a.xlsx b.csv -o out_dir/
  預設輸出資料夾：./output/cantab_renamed/

需求套件：pandas、openpyxl
"""

import os
import sys
import argparse
from collections import OrderedDict

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.utils.wait_and_retry import wait_and_retry
from cantab_config import (
    load_rename_map,
    load_field_rules,
    CANTAB_RENAME_XLSX_NAME,
    HEADER_RENAMES,
)

# 讓中文/emoji 在 Windows 主控台（cp950）也能正常輸出
try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

# ── CONFIG ──────────────────────────────────────────────────────────────
from datetime import date

TODAY = date.today().strftime("%Y%m%d")
DEFAULT_OUTPUT_DIR = os.path.join(".", "output", "cantab_renamed")
REPORT_NAME = f"{TODAY}_CANTAB_rename_report.xlsx"

CAT_CASE_FIX = "db_code 大小寫/空白修正"
CAT_NO_CODE = "明定不收（丟棄）"
CAT_UNKNOWN = "未知欄位（保留原名，人工比對）"
# ────────────────────────────────────────────────────────────────────────

# ---- 報表樣式（沿用 cantab_precheck.py 的風格）----
FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="4472C4")
WARN_FILL = PatternFill("solid", start_color="FFF2CC")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=12)
BASE_FONT = Font(name=FONT_NAME)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── 報表工具（沿用 cantab_precheck.py）──────────────────────────────────
def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autofit(ws, max_width=50):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 0
        for cell in col:
            if cell.value is not None:
                longest = max(longest, max(len(s) for s in str(cell.value).split("\n")))
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), max_width)


def write_table(ws, headers, rows, start_row=1, warn_col=None, warn_value=None):
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h).font = BASE_FONT
    style_header(ws, len(headers), row=start_row)
    for i, rowvals in enumerate(rows, start=start_row + 1):
        flag = warn_col is not None and rowvals[warn_col] == warn_value
        for j, v in enumerate(rowvals, start=1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = BASE_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if flag:
                cell.fill = WARN_FILL
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


# ── 檔案蒐集 ────────────────────────────────────────────────────────────
def collect_jobs(paths):
    """回傳 [(輸出檔基底名, [檔案路徑, ...]), ...]。

    資料夾 → 一個 job（該夾所有檔案彙整成一個輸出，基底名 = 資料夾名）；
    單一檔案 → 一檔一 job（基底名 = 去副檔名的檔名）。
    """
    jobs = []
    for p in paths:
        if os.path.isdir(p):
            files = [
                os.path.join(p, f) for f in sorted(os.listdir(p))
                if f.endswith((".xlsx", ".xls", ".csv")) and not f.startswith("~$")
            ]
            if not files:
                print(f"⚠️  略過：{p} 內沒有 xlsx/csv 檔案")
                continue
            jobs.append((os.path.basename(os.path.normpath(p)), files))
        elif os.path.isfile(p):
            jobs.append((os.path.splitext(os.path.basename(p))[0], [p]))
        else:
            print(f"⚠️  略過：找不到 {p}")
    return jobs


def read_sheets(filepath):
    """讀取一個檔案的所有工作表，回傳 {工作表名: DataFrame}（值一律 str）。

    csv 編碼 fallback：utf-8-sig → cp950。
    """
    if filepath.lower().endswith(".csv"):
        last_err = None
        for enc in ("utf-8-sig", "cp950"):
            try:
                return {"Sheet1": pd.read_csv(filepath, dtype=str, encoding=enc)}
            except UnicodeDecodeError as e:
                last_err = e
        raise last_err
    return pd.read_excel(filepath, sheet_name=None, dtype=str)


# ── 核心改名 ────────────────────────────────────────────────────────────
def rename_sheet(df, rename_map, no_code_set, code_lookup):
    """對單一工作表分類欄位並改名。依序判斷：
      1. 已是 db_code（code_lookup：小寫 → 標準寫法）→ 直接使用（必要時修正大小寫/空白）
      2. 對照表 raw 欄名（rename_map）→ 改名成 new_db_code
      3. 明定不收（no_code_set）→ 丟棄
      4. 未知欄位 → 保留原名，供人工比對

    回傳 (renamed_df, review, stats, error)：
      review：[(原始欄名, 分類)]，分類 ∈ {CAT_CASE_FIX, CAT_NO_CODE, CAT_UNKNOWN}
      stats：{"renamed", "direct", "no_code", "unknown"} 各分類欄數
      error：同一工作表兩個來源欄位映射到同一目標欄名時的錯誤訊息，否則 None
    """
    keep = OrderedDict()   # 原始欄名 -> 新欄名
    review = []
    stats = {"renamed": 0, "direct": 0, "no_code": 0, "unknown": 0}
    target_sources = {}    # 新欄名 -> 原始欄名（防呆用）

    def claim(target, col):
        if target in target_sources:
            return (f"欄位 {target_sources[target]!r} 與 {col!r} "
                    f"皆映射到 {target!r}")
        target_sources[target] = col
        keep[col] = target
        return None

    for col in df.columns:
        key = str(col).strip().lower() if isinstance(col, str) else ""
        if key in code_lookup:
            target = code_lookup[key]
            err = claim(target, col)
            if err:
                return None, review, stats, err
            stats["direct"] += 1
            if str(col) != target:
                review.append((str(col), CAT_CASE_FIX))
        elif key in rename_map:
            err = claim(rename_map[key], col)
            if err:
                return None, review, stats, err
            stats["renamed"] += 1
        elif key in no_code_set:
            stats["no_code"] += 1
            review.append((str(col), CAT_NO_CODE))
        else:
            err = claim(col, col)
            if err:
                return None, review, stats, err
            stats["unknown"] += 1
            review.append((str(col), CAT_UNKNOWN))

    renamed = df[list(keep.keys())].rename(columns=keep)
    return renamed, review, stats, None


# ── 報告輸出 ────────────────────────────────────────────────────────────
def build_report(output_path, file_stats, review_rows):
    wb = Workbook()

    # ---- Summary ----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "CANTAB 欄位改名報告"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (f"對照表：{CANTAB_RENAME_XLSX_NAME}｜"
                f"未知欄位保留原名、整列標黃，需人工比對")
    ws["A2"].font = BASE_FONT
    write_table(
        ws,
        ["檔案（｜工作表）", "改名", "db_code直用", "丟棄(明定不收)", "未知保留", "狀態"],
        file_stats if file_stats else [("（無）", "", "", "", "", "")],
        start_row=4,
    )
    autofit(ws, max_width=60)

    # ---- Review（僅列需要注意的欄位；一般改名不列）----
    ws = wb.create_sheet("Review")
    if not review_rows:
        review_rows = [("（無）", "", "沒有需要注意的欄位")]
    write_table(
        ws,
        ["檔案（｜工作表）", "原始欄名", "分類"],
        review_rows, warn_col=2, warn_value=CAT_UNKNOWN,
    )
    autofit(ws, max_width=60)

    wait_and_retry(lambda: wb.save(output_path), output_path)


def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="CANTAB 原始資料欄位改名：raw 欄名 → new_db_code（依對照表）")
    parser.add_argument("paths", nargs="+", help="要改名的 xlsx/csv 檔案或資料夾")
    parser.add_argument("-o", "--output-dir", default=DEFAULT_OUTPUT_DIR,
                        metavar="DIR",
                        help=f"輸出資料夾（預設 {DEFAULT_OUTPUT_DIR}）")
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)

    jobs = collect_jobs(args.paths)
    if not jobs:
        print("Error: 找不到任何 xlsx/csv 檔案")
        sys.exit(1)

    rename_map, no_code_set = load_rename_map()
    field_rules = load_field_rules()
    codes = set(rename_map.values())
    in_json = codes & set(field_rules)
    # 已是 db_code 的欄名直接使用：field.json 欄位 + 對照表 new_db_code（小寫 → 標準寫法）
    code_lookup = {c.strip().lower(): c for c in (set(field_rules) | codes)}
    n_files = sum(len(f) for _, f in jobs)
    print(f"📄 檔案數：{n_files}（輸出 {len(jobs)} 檔）｜輸出資料夾：{args.output_dir}")
    print(f"📐 對照表：{CANTAB_RENAME_XLSX_NAME}｜"
          f"改名對應 {len(rename_map)} 組（其中 {len(in_json)} 個目標欄名"
          f"存在於 field.json）｜明定不收 {len(no_code_set)} 欄｜"
          f"可辨識 db_code {len(code_lookup)} 個")
    # header 目標欄名（raw_cantab_age、nart）本來就不在 field.json，不算異常
    not_in_json = sorted(codes - set(field_rules) - set(HEADER_RENAMES.values()))
    if not_in_json:
        print(f"⚠️  目標欄名不在 field.json 中：{', '.join(not_in_json)}")

    os.makedirs(args.output_dir, exist_ok=True)

    file_stats = []     # (來源, 改名數, 直用數, 明定不收數, 未知保留數, 狀態)
    review_rows = []    # (來源, 原始欄名, 分類)
    n_ok = 0
    any_unknown = False

    for out_base, files in jobs:
        print(f"\n處理: {out_base}（{len(files)} 檔）")
        frames = []
        n_skipped = 0

        for filepath in files:
            filename = os.path.basename(filepath)
            try:
                sheets = read_sheets(filepath)
            except Exception as e:
                print(f"  ✗ {filename} 無法讀取: {e}")
                file_stats.append((filename, 0, 0, 0, 0, f"讀取失敗：{e}"))
                n_skipped += 1
                continue

            multi = len(sheets) > 1
            file_frames = []
            file_error = None
            for sheet_name, df in sheets.items():
                source = f"{filename}｜{sheet_name}" if multi else filename
                renamed, review, stats, error = rename_sheet(
                    df, rename_map, no_code_set, code_lookup)
                review_rows.extend((source, col, cat) for col, cat in review)

                if error:
                    file_error = f"{source}：{error}"
                    file_stats.append((source, stats["renamed"], stats["direct"],
                                       stats["no_code"], stats["unknown"],
                                       f"錯誤：{error}"))
                    break

                file_frames.append(renamed)
                file_stats.append((source, stats["renamed"], stats["direct"],
                                   stats["no_code"], stats["unknown"], "OK"))
                unknown_note = (f"｜⚠️ 未知保留 {stats['unknown']} 欄"
                                if stats["unknown"] else "")
                if stats["unknown"]:
                    any_unknown = True
                print(f"  [{source}] 改名 {stats['renamed']}｜"
                      f"db_code直用 {stats['direct']}｜"
                      f"丟棄(明定不收) {stats['no_code']}{unknown_note}")

            if file_error:
                print(f"  ✗ {file_error}（此檔不列入彙整）")
                n_skipped += 1
                continue

            frames.extend(file_frames)

        if not frames:
            print(f"  ✗ {out_base}：沒有可輸出的資料")
            continue

        merged = pd.concat(frames, ignore_index=True, sort=False)
        out_path = os.path.join(args.output_dir, out_base + ".xlsx")
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            merged.to_excel(writer, sheet_name="Sheet1", index=False)
        n_ok += 1
        skipped_note = f"｜略過 {n_skipped} 檔" if n_skipped else ""
        print(f"  → {out_base}.xlsx：{len(merged)} 列 × {len(merged.columns)} 欄"
              f"（彙整 {len(files) - n_skipped} 檔{skipped_note}）")

    report_path = os.path.join(args.output_dir, REPORT_NAME)
    build_report(report_path, file_stats, review_rows)

    print("\n===== 改名結果 =====")
    print(f"成功輸出 {n_ok}/{len(jobs)} 檔｜輸出資料夾：{args.output_dir}")
    if any_unknown:
        print("⚠️  有未知欄位以原名保留，請檢視報告 Review 分頁（標黃列）並人工比對")
    print(f"已輸出報告：{report_path}")


if __name__ == "__main__":
    main()
