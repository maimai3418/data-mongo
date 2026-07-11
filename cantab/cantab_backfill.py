#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CANTAB datapool 重複確認 + session_start_time 回填嘗試。

步驟一（重複確認，概念同 CPT hard dedup）：
  以 field.json 中除了 birth_date, sex, record_date 以外的欄位 + raw_cantab_age
  作為比對欄位，確認 datapool 內沒有完全重複的資料列。

步驟二（session_start_time 回填嘗試）：
  對缺 session_start_time 的列，蒐集同 famid 的「錨點日期」：
    - datapool 內其他有 session_start_time 的列
    - DB CANTAB collection 中有 record_date 的既有紀錄
  以 Participants birth_date + raw_cantab_age 驗證錨點（年齡差 ≤ AGE_TOL），
  通過驗證的錨點日期依 WINDOW_DAYS 聚類（同 visit）：
    - anchor_single       僅一個 cluster → 提議回填該 cluster 代表日期（最早日）
    - anchor_ambiguous    ≥2 個 cluster → 匯出候選日期供人工判定
    - anchor_no_age_match 有錨點但年齡皆不符 → 僅提供估算日期
    - anchor_unvalidated  有錨點但缺 birth_date 或 age，無法驗證 → 人工判定
    - age_estimate_only   無錨點，僅能以 birth_date + age 估算（依 CLAUDE.md
                          慣例不直接回填，僅列於 estimated_date 欄供參考）
    - no_backfill         無任何線索
  本腳本不改動原始檔、不寫入 DB，僅輸出報表（Records 工作表含提議回填欄位，
  原始 session_start_time 欄位保留原值）。

用法（從專案根目錄執行）：
  python cantab/cantab_backfill.py [datapool.xlsx] [-o report.xlsx]

需求套件：pandas、openpyxl、pymongo、python-dotenv
"""

import os
import sys
import json
import argparse
from datetime import datetime, timedelta

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

import pandas as pd
from openpyxl import Workbook

from dotenv import load_dotenv
load_dotenv()

from src.importer import get_db
from src.age_matcher import parse_date, calculate_age
from src.utils.wait_and_retry import wait_and_retry
from cantab_config import CANTAB_COLLECTION, RAW_AGE_FIELD, discover_field_json
from cantab_precheck import (
    norm_value, norm_key, fmt_sources, write_table, autofit,
    style_header, BASE_FONT, TITLE_FONT,
)

# 讓中文/emoji 在 Windows 主控台（cp950）也能正常輸出
try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

# === CONFIG ===
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TODAY = datetime.now().strftime("%Y%m%d")
DEFAULT_INPUT = os.path.join(SCRIPT_DIR, "20260710_cantab_datapool_all.xlsx")
DEFAULT_OUTPUT = os.path.join(SCRIPT_DIR, f"{TODAY}_CANTAB_sst_backfill.xlsx")
COLLECTION = CANTAB_COLLECTION
SST_FIELD = "session_start_time"
# 重複比對：field.json 全欄位排除下列三欄，另加 raw_cantab_age
EXCLUDE_DUP_FIELDS = {"birth_date", "sex", "record_date"}
AGE_TOL = 1.0     # 錨點年齡驗證容差（年），同 src/age_matcher.find_age_conflicts
WINDOW_DAYS = 14  # for date clustering（同 visit 判定）
# === END CONFIG ===

STATUS_LABELS = {
    "anchor_single": "anchor_single（提議回填）",
    "anchor_ambiguous": "anchor_ambiguous（人工判定）",
    "anchor_no_age_match": "anchor_no_age_match（錨點年齡不符）",
    "anchor_unvalidated": "anchor_unvalidated（無法驗證）",
    "age_estimate_only": "age_estimate_only（僅估算）",
    "no_backfill": "no_backfill（無線索）",
}


# ── 基礎工具 ────────────────────────────────────────────────────────────
def clean_famid(v):
    """famid 一律字串處理：strip + 去 Excel 浮點殘留 .0。"""
    n = norm_key(v)
    return n[:-2] if n.endswith(".0") else n


def norm_date(v):
    """任意日期表示 → YYYY-MM-DD 文字；無法解析 → None。"""
    n = norm_value(v)
    if n is None:
        return None
    ts = pd.to_datetime(str(n), errors="coerce")
    return None if pd.isna(ts) else ts.strftime("%Y-%m-%d")


def cluster_dates(date_strs, window_days=WINDOW_DAYS):
    """日期字串聚類：與 cluster 起始日相距 ≤ window_days 視為同一 visit。

    回傳 [[datetime, ...], ...]（各 cluster 內由早到晚）。
    """
    ds = sorted(pd.to_datetime(d) for d in set(date_strs))
    clusters = []
    for d in ds:
        if clusters and (d - clusters[-1][0]).days <= window_days:
            clusters[-1].append(d)
        else:
            clusters.append([d])
    return clusters


def estimate_date(birth_date, raw_age):
    """birth_date + age（年）估算施測日期（同 src/age_matcher.estimate_record_date，
    但用預先批次撈好的 birth_date，不逐列查 DB）。"""
    b = parse_date(birth_date)
    if not b or raw_age is None:
        return None
    return (b + timedelta(days=raw_age * 365.25)).strftime("%Y-%m-%d")


# ── 步驟一：重複確認 ────────────────────────────────────────────────────
def load_dup_check_fields():
    """回傳 (json 路徑, 比對欄位清單)：field.json 欄位（原始大小寫）
    排除 EXCLUDE_DUP_FIELDS，另加 raw_cantab_age。"""
    path = discover_field_json()
    if not path:
        raise SystemExit("Error: 找不到 CANTAB 的 fields.json")
    with open(path, "r", encoding="utf-8") as f:
        fields = json.load(f)
    check = [c for c in fields if c not in EXCLUDE_DUP_FIELDS]
    check.append(RAW_AGE_FIELD)
    return path, check


def find_duplicates(df, check_fields):
    """回傳 (實際比對欄位, 檔內缺少的比對欄位, dup_groups)。

    dup_groups：[(代表列 dict, [Excel 列號, ...]), ...]，完全重複的列群組。
    """
    present = [c for c in check_fields if c in df.columns]
    missing = [c for c in check_fields if c not in df.columns]

    normed = df[present].map(norm_value)
    sig = normed.apply(tuple, axis=1)
    dup_mask = sig.duplicated(keep=False)

    dup_groups = []
    if dup_mask.any():
        for _, idxs in sig[dup_mask].groupby(sig[dup_mask]).groups.items():
            rows = [int(i) + 2 for i in idxs]  # +1 標題列、+1 轉 1-based
            dup_groups.append((df.loc[idxs[0]].to_dict(), rows))
    return present, missing, dup_groups


# ── 步驟二：回填分類 ────────────────────────────────────────────────────
def build_anchor_maps(df, db, famids):
    """回傳 (anchor_map, birth_map)。

    anchor_map：{famid: {日期: 來源集合}}，來源為 'datapool' / 'db'。
    birth_map ：{famid: birth_date 字串}（Participants）。
    """
    anchor_map = {}

    def add(famid, date_str, source):
        if famid and date_str:
            anchor_map.setdefault(famid, {}).setdefault(date_str, set()).add(source)

    for _, row in df.iterrows():
        add(clean_famid(row.get("famid")), norm_date(row.get(SST_FIELD)), "datapool")

    for doc in db[COLLECTION].find(
            {"famid": {"$in": famids}, "record_date": {"$nin": [None, ""]}},
            {"famid": 1, "record_date": 1}):
        add(clean_famid(doc["famid"]), norm_date(doc["record_date"]), "db")

    birth_map = {}
    for doc in db["Participants"].find(
            {"famid": {"$in": famids}, "birth_date": {"$nin": [None, ""]}},
            {"famid": 1, "birth_date": 1}):
        birth_map[clean_famid(doc["famid"])] = doc["birth_date"]

    return anchor_map, birth_map


def classify_row(famid, raw_age, anchor_map, birth_map):
    """對一列缺 session_start_time 的資料分類，回傳 dict：
    status / backfill / sources / anchor_dates / matched_dates / estimated / note
    """
    anchors = anchor_map.get(famid, {})
    birth = birth_map.get(famid)
    estimated = estimate_date(birth, raw_age) if birth else None

    res = {
        "status": None, "backfill": None, "sources": "",
        "anchor_dates": "", "matched_dates": "", "estimated": estimated,
        "note": "",
    }
    if anchors:
        res["anchor_dates"] = ", ".join(
            f"{d}({'/'.join(sorted(srcs))})" for d, srcs in sorted(anchors.items()))

    if not anchors:
        if estimated:
            res["status"] = "age_estimate_only"
        else:
            res["status"] = "no_backfill"
            res["note"] = ("Participants 中找不到此 famid 或無生日資料"
                           if raw_age is not None else "缺 raw_cantab_age")
        return res

    if not birth or raw_age is None:
        res["status"] = "anchor_unvalidated"
        res["note"] = "缺 birth_date" if not birth else "缺 raw_cantab_age"
        return res

    matched = [d for d in anchors
               if (lambda a: a is not None and abs(a - raw_age) <= AGE_TOL)
                  (calculate_age(birth, d))]
    res["matched_dates"] = ", ".join(sorted(matched))

    if not matched:
        res["status"] = "anchor_no_age_match"
        return res

    clusters = cluster_dates(matched)
    if len(clusters) == 1:
        rep = clusters[0][0].strftime("%Y-%m-%d")
        res["status"] = "anchor_single"
        res["backfill"] = rep
        res["sources"] = "/".join(sorted(anchors[rep])) if rep in anchors else "cluster"
    else:
        res["status"] = "anchor_ambiguous"
        res["note"] = f"{len(clusters)} 個 visit cluster"
    return res


# ── 報表輸出 ────────────────────────────────────────────────────────────
def build_report(output_path, input_path, json_path, present, missing_fields,
                 dup_groups, df, results, status_counts):
    wb = Workbook()

    # ---- Summary ----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "CANTAB datapool 重複確認 + session_start_time 回填嘗試"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (f"輸入：{os.path.basename(input_path)}｜"
                f"比對欄位：{len(present)} 個（{os.path.basename(json_path)} "
                f"排除 {', '.join(sorted(EXCLUDE_DUP_FIELDS))}，加 {RAW_AGE_FIELD}）｜"
                f"AGE_TOL={AGE_TOL} 年｜WINDOW_DAYS={WINDOW_DAYS} 天")
    ws["A2"].font = BASE_FONT

    n_sst = int(df[SST_FIELD].map(norm_value).notna().sum())
    summary_rows = [
        ("總列數", len(df)),
        ("比對欄位（檔內缺少）", ", ".join(missing_fields) if missing_fields else "（無）"),
        ("完全重複群組數", len(dup_groups)),
        ("完全重複列數", sum(len(r) for _, r in dup_groups)),
        ("已有 session_start_time 列數", n_sst),
        ("缺 session_start_time 列數", len(results)),
    ]
    summary_rows += [(STATUS_LABELS[s], status_counts.get(s, 0))
                     for s in STATUS_LABELS]
    write_table(ws, ["項目", "數值"], summary_rows, start_row=4)
    autofit(ws, max_width=80)

    # ---- Duplicates ----
    ws = wb.create_sheet("Duplicates")
    dup_rows = []
    for rep, rows in dup_groups:
        dup_rows.append((clean_famid(rep.get("famid")),
                         norm_key(rep.get(RAW_AGE_FIELD)),
                         len(rows), ", ".join(f"列{r}" for r in rows)))
    if not dup_rows:
        dup_rows = [("（無）", "", "", "沒有完全重複的資料列")]
    write_table(ws, ["famid", RAW_AGE_FIELD, "重複筆數", "出現列號"], dup_rows)
    autofit(ws, max_width=70)

    # ---- Backfill（缺 sst 列的分類明細）----
    ws = wb.create_sheet("Backfill")
    bf_rows = []
    for r in results:
        bf_rows.append((
            r["excel_row"], r["famid"], r["raw_age"], r["birth_date"],
            r["status"], r["backfill"], r["sources"],
            r["anchor_dates"], r["matched_dates"], r["estimated"], r["note"],
        ))
    if not bf_rows:
        bf_rows = [("（無）", *[""] * 10)]
    write_table(
        ws,
        ["Excel列", "famid", RAW_AGE_FIELD, "birth_date", "狀態",
         "提議回填 session_start_time", "回填來源", "錨點日期(全部)",
         "年齡相符錨點", "估算日期(birth+age)", "備註"],
        bf_rows, warn_col=4, warn_value="anchor_ambiguous",
    )
    autofit(ws, max_width=60)

    # ---- Records（完整資料 + 提議回填欄位；原始欄位保留原值）----
    ws = wb.create_sheet("Records")
    extra_cols = ["sst_backfill_status", "sst_backfill", "sst_backfill_source",
                  "sst_estimated_date"]
    headers = list(df.columns) + extra_cols
    ws.append(headers)
    style_header(ws, len(headers))

    by_index = {r["index"]: r for r in results}
    for idx, row in df.iterrows():
        r = by_index.get(idx)
        extra = (["has_sst", None, None, None] if r is None else
                 [r["status"], r["backfill"], r["sources"] or None, r["estimated"]])
        ws.append([norm_value(v) for v in row.tolist()] + extra)
    ws.freeze_panes = "A2"

    wait_and_retry(lambda: wb.save(output_path), output_path)
    return summary_rows


def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="CANTAB datapool 重複確認 + session_start_time 回填嘗試")
    parser.add_argument("input", nargs="?", default=DEFAULT_INPUT,
                        help=f"datapool xlsx（預設 {DEFAULT_INPUT}）")
    parser.add_argument("-o", "--output", default=DEFAULT_OUTPUT, metavar="XLSX",
                        help=f"報表輸出路徑（預設 {DEFAULT_OUTPUT}）")
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)
    if not os.path.isfile(args.input):
        raise SystemExit(f"Error: 找不到輸入檔 {args.input}")

    df = pd.read_excel(args.input, dtype=str)
    df = df.dropna(how="all").reset_index(drop=True)
    print(f"📄 輸入：{args.input}（{len(df)} 列）")

    # ---- 步驟一：重複確認 ----
    json_path, check_fields = load_dup_check_fields()
    present, missing_fields, dup_groups = find_duplicates(df, check_fields)
    print(f"📐 重複比對欄位：{len(present)} 個（來源 {os.path.basename(json_path)}）")
    if missing_fields:
        print(f"⚠️  檔內缺少比對欄位：{', '.join(missing_fields)}")
    if dup_groups:
        print(f"⚠️  發現完全重複 {len(dup_groups)} 組"
              f"（共 {sum(len(r) for _, r in dup_groups)} 列），詳見 Duplicates 工作表")
    else:
        print("✅ 無完全重複的資料列")

    # ---- 步驟二：session_start_time 回填 ----
    db = get_db()
    famids = sorted({clean_famid(v) for v in df["famid"]} - {""})
    anchor_map, birth_map = build_anchor_maps(df, db, famids)
    print(f"🔎 錨點：datapool+DB 共 {sum(len(v) for v in anchor_map.values())} 個日期"
          f"（{len(anchor_map)} 個 famid）｜Participants birth_date {len(birth_map)} 筆")

    results = []
    for idx, row in df.iterrows():
        if norm_value(row.get(SST_FIELD)) is not None:
            continue
        famid = clean_famid(row.get("famid"))
        raw_age = norm_value(row.get(RAW_AGE_FIELD))
        raw_age = float(raw_age) if isinstance(raw_age, (int, float)) else None
        res = classify_row(famid, raw_age, anchor_map, birth_map)
        res.update({
            "index": idx, "excel_row": idx + 2, "famid": famid,
            "raw_age": raw_age, "birth_date": birth_map.get(famid),
        })
        results.append(res)

    status_counts = {}
    for r in results:
        status_counts[r["status"]] = status_counts.get(r["status"], 0) + 1

    summary_rows = build_report(
        args.output, args.input, json_path, present, missing_fields,
        dup_groups, df, results, status_counts)

    print("\n===== 檢查結果 =====")
    for label, value in summary_rows:
        print(f"{label}：{value}")
    print(f"\n已輸出報表：{args.output}")


if __name__ == "__main__":
    main()
