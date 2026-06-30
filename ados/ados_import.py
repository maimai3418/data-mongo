#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ADOS 匯入：讀取已重命名的 ADOS xlsx（含 M1/M2/M3/M4 工作表），
依工作表名稱匯入對應 collection（M1→ADOS_M1、M2→ADOS_M2…）。

與一般 scale 匯入（main.py）邏輯類似：
  - 相同 famid + record_date 視為重複，不重複匯入（$setOnInsert；**不看 role**）。
  - item 欄位全為空的列略過不上傳。
數值轉換不同：只有等於 999 才轉 null，其餘按原始值上傳（0 傳 0、1 傳 1…）。
欄位限制讀自 ados_config.py（待補；補上後自動生效）。

錯誤匯出至 {today}_ados_import_error.xlsx。

用法（從專案根目錄執行）：
  python ados/ados_import.py <檔案或資料夾> [更多路徑...]
  python ados/ados_import.py renamed_dir/                 # 資料夾內所有 xlsx
  python ados/ados_import.py a.xlsx b.xlsx
  python ados/ados_import.py renamed_dir/ --modules M1 M3  # 只匯入指定 module
  python ados/ados_import.py renamed_dir/ --dry-run        # 只驗證、不寫入 DB
  python ados/ados_import.py a.xlsx --project-code NHRI113 # 順便標註計畫代碼
  python ados/ados_import.py a.xlsx -e errors.xlsx         # 自訂錯誤輸出檔名

需求套件：pandas、openpyxl、pymongo、python-dotenv
"""

import os
import sys
import argparse
import warnings
from datetime import date

# 讓本腳本可從 ados/ 子目錄被直接執行：把專案根目錄加入 sys.path，
# 以便 import 根目錄的 src 套件（get_db 等）。
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

import pandas as pd
from dotenv import load_dotenv
from pymongo import UpdateOne
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.importer import get_db
from ados_config import (
    ADOS_SHARED_FIELDS,
    ADOS_COLLECTION_MAP,
    ADOS_NULL_VALUE,
    ADOS_ITEM_PREFIX,
    ADOS_FIELDS_DIR,
    load_field_rules,
)

# 讓中文/emoji 在 Windows 主控台（cp950）也能正常輸出
try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

load_dotenv()

TODAY = date.today().strftime("%Y%m%d")
DEFAULT_ERROR_OUTPUT = f"{TODAY}_ados_import_error.xlsx"
PROJECT_FIELD = "research_project_code"


# ── 數值處理 ────────────────────────────────────────────────────────────
def is_blank(val):
    """空值判定：None 或 NaN（dtype=str 的空白儲存格會是 float nan）。"""
    return val is None or (isinstance(val, float) and pd.isna(val))


def to_number(s):
    """把字串轉成 int（整數值）或 float；無法轉則回 None。"""
    try:
        f = float(s)
    except (ValueError, TypeError):
        return None
    return int(f) if f.is_integer() else f


def clean_value(val):
    """header 欄位：去前後空白、空字串→None，其餘維持原始字串。"""
    if is_blank(val):
        return None
    s = str(val).strip()
    return s if s != "" else None


def to_date_text(col, val):
    """日期欄位：統一轉成 YYYY-MM-DD 文字。

    空值→None；可解析→YYYY-MM-DD；無法解析→保留原值並回報錯誤（該列會被拒）。
    """
    if is_blank(val):
        return None, None
    s = str(val).strip()
    if s == "":
        return None, None
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")  # 忽略 dayfirst 等格式推斷警告
        dt = pd.to_datetime(s, errors="coerce")
    if pd.isna(dt):
        return s, f"{col} 日期格式錯誤"
    return dt.strftime("%Y-%m-%d"), None


def get_item_rule(field_rules, module, col):
    """從 field.json 解析出的規則表取得該 item 欄位規則（無規則回 None）。"""
    return field_rules.get(module, {}).get(col)


def convert_item(col, val, rule):
    """item 欄位轉換，回傳 (converted, error)。

    規則：空值→null；等於 999→null；其餘按原始值（能轉數字就存數字）。
    若該欄位設有 rule（int/float + range），則額外做型別與範圍驗證。
    """
    if is_blank(val):
        return None, None
    s = str(val).strip()
    if s == "":
        return None, None

    num = to_number(s)
    # 只有 999 轉 null
    if num is not None and num == ADOS_NULL_VALUE:
        return None, None

    if rule is None:
        # 不驗證：原始值上傳（能轉數字就用數字，否則保留原字串）
        return (num if num is not None else s), None

    rtype = rule.get("type")
    if rtype in ("int", "float"):
        if num is None:
            return None, f"{col} 數值錯誤"
        if rtype == "int" and not isinstance(num, int):
            return None, f"{col} 數值錯誤(非整數)"
        if "range" in rule:
            lo, hi = rule["range"]
            if not (lo <= num <= hi):
                return None, f"{col} 數值錯誤"
        return num, None

    # 其他型別：原樣上傳
    return (num if num is not None else s), None


# ── 單一工作表處理 ──────────────────────────────────────────────────────
def process_sheet(filename, module, df, errors, field_rules):
    """處理單一 module 工作表，回傳 (docs, n_empty)。錯誤累加到 errors。"""
    df = df.copy()
    df.columns = df.columns.str.strip()
    df = df.loc[:, df.columns.notna()]
    df = df.dropna(how="all")

    item_prefix = ADOS_ITEM_PREFIX.format(module=module.lower())
    item_cols = [c for c in df.columns if c.startswith(item_prefix)]
    other_cols = [c for c in df.columns if c not in item_cols]

    docs = []
    n_empty = 0

    for _, row in df.iterrows():
        rowd = row.to_dict()
        doc = {}
        row_errors = []

        # header / 其他欄位：<DATE> 轉 YYYY-MM-DD 文字，其餘原樣（清空白）
        for col in other_cols:
            rule = field_rules.get(module, {}).get(col)
            if rule and rule.get("type") == "date":
                new_val, err = to_date_text(col, rowd.get(col))
                if err:
                    row_errors.append(err)
                doc[col] = new_val
            else:
                doc[col] = clean_value(rowd.get(col))

        # item 欄位：999→null + 規則驗證（規則來自 field.json）
        for col in item_cols:
            new_val, err = convert_item(col, rowd.get(col), get_item_rule(field_rules, module, col))
            if err:
                row_errors.append(err)
            else:
                doc[col] = new_val

        # item 欄位全為空 → 略過不上傳
        if item_cols and all(doc.get(c) is None for c in item_cols):
            n_empty += 1
            continue

        # 必要欄位檢查（famid / record_date）
        missing = [
            f for f in ADOS_SHARED_FIELDS
            if doc.get(f) is None or str(doc.get(f)).strip() == ""
        ]
        if missing:
            row_errors.append(f"缺少必要欄位: {', '.join(missing)}")

        if row_errors:
            errors.append({
                "file": filename,
                "module": module,
                "famid": clean_value(rowd.get("famid")),
                "record_date": clean_value(rowd.get("record_date")),
                "error": "; ".join(row_errors),
            })
        else:
            docs.append(doc)

    return docs, n_empty


# ── 寫入 DB ─────────────────────────────────────────────────────────────
def upsert_ados(db, collection_name, docs):
    """以 famid + record_date 為 unique key 上傳；重複者略過。回傳 (inserted, skipped)。"""
    if not docs:
        return 0, 0
    operations = [
        UpdateOne(
            {k: doc.get(k) for k in ADOS_SHARED_FIELDS},
            {"$setOnInsert": doc},
            upsert=True,
        )
        for doc in docs
    ]
    result = db[collection_name].bulk_write(operations)
    inserted = result.upserted_count
    skipped = len(docs) - inserted
    return inserted, skipped


# ── 錯誤匯出 ────────────────────────────────────────────────────────────
def export_errors(errors, output_path):
    """匯出錯誤報告至 Excel。"""
    if not errors:
        print("\n✓ 無錯誤，不產生 error 檔案")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "errors"

    headers = ["file", "module", "famid", "record_date", "error"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="C0392B")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, err in enumerate(errors, 2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=err.get(key))

    widths = {"A": 36, "B": 8, "C": 14, "D": 14, "E": 60}
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"\n✗ 共 {len(errors)} 筆錯誤，已匯出: {os.path.basename(output_path)}")


# ── 檔案蒐集 ────────────────────────────────────────────────────────────
def collect_files(paths):
    """把 path 清單（檔案或資料夾）展開成 xlsx 檔案清單（排序、去重）。"""
    files = []
    for p in paths:
        if os.path.isdir(p):
            for f in sorted(os.listdir(p)):
                if (f.endswith(".xlsx") and not f.startswith("~$")
                        and not f.startswith(TODAY)):
                    files.append(os.path.join(p, f))
        elif os.path.isfile(p):
            files.append(p)
        else:
            print(f"⚠️  略過：找不到 {p}")
    # 去重但保留順序
    seen, unique = set(), []
    for f in files:
        ap = os.path.abspath(f)
        if ap not in seen:
            seen.add(ap)
            unique.append(f)
    return unique


def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="ADOS 匯入：依工作表名稱（M1~M4）匯入對應 collection（ADOS_M1~M4）")
    parser.add_argument("paths", nargs="+",
                        help="要匯入的 xlsx 檔案或資料夾（資料夾會展開其下所有 xlsx）")
    parser.add_argument("--modules", nargs="+", default=list(ADOS_COLLECTION_MAP.keys()),
                        metavar="M", choices=list(ADOS_COLLECTION_MAP.keys()),
                        help=f"只匯入指定 module（預設全部：{', '.join(ADOS_COLLECTION_MAP)}）")
    parser.add_argument("--project-code", default=None, metavar="CODE",
                        help="替每筆資料標註計畫代碼（寫入 research_project_code，選填）")
    parser.add_argument("--dry-run", action="store_true",
                        help="只驗證與轉換、不寫入 MongoDB")
    parser.add_argument("-e", "--error-output", default=DEFAULT_ERROR_OUTPUT, metavar="XLSX",
                        help=f"錯誤報告輸出路徑（預設 {DEFAULT_ERROR_OUTPUT}）")
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)

    files = collect_files(args.paths)
    if not files:
        print("Error: 找不到任何 xlsx 檔案")
        sys.exit(1)

    modules = args.modules
    project_code = (args.project_code or "").strip() or None

    # 欄位限制直接讀自 ados/fields 的 field.json
    field_rules = load_field_rules()

    print(f"📄 檔案數：{len(files)}｜匯入 module：{', '.join(modules)}"
          f"{'｜DRY-RUN（不寫入）' if args.dry_run else ''}")
    print(f"📐 欄位限制來源：{ADOS_FIELDS_DIR}")
    for m in modules:
        n_ranged = sum(1 for r in field_rules.get(m, {}).values() if r and "range" in r)
        print(f"   [{m}] 有範圍限制的欄位：{n_ranged}")
    if project_code:
        print(f"🏷️  計畫代碼：{project_code}")

    errors = []
    # 每個 module 累積所有檔案的 docs，最後一次 upsert
    docs_by_module = {m: [] for m in modules}
    empty_by_module = {m: 0 for m in modules}

    for filepath in files:
        filename = os.path.basename(filepath)
        print(f"\n處理: {filename}")
        try:
            all_sheets = pd.read_excel(filepath, sheet_name=None, dtype=str)
        except Exception as e:
            errors.append({"file": filename, "module": "-", "famid": "",
                           "record_date": "", "error": f"無法讀取檔案: {e}"})
            continue

        for module in modules:
            if module not in all_sheets:
                print(f"  [{module}] 工作表不存在，略過")
                continue
            docs, n_empty = process_sheet(filename, module, all_sheets[module], errors, field_rules)
            if project_code:
                for doc in docs:
                    doc[PROJECT_FIELD] = project_code
            docs_by_module[module].extend(docs)
            empty_by_module[module] += n_empty
            print(f"  [{module}] 有效 {len(docs)}｜略過(空值) {n_empty}")

    # ── 寫入 MongoDB ──
    print("\n===== 匯入結果 =====")
    db = None if args.dry_run else get_db()
    for module in modules:
        collection = ADOS_COLLECTION_MAP[module]
        docs = docs_by_module[module]
        n_err = len([e for e in errors if e.get("module") == module])
        if args.dry_run:
            print(f"[{collection}] dry-run：待上傳 {len(docs)}｜"
                  f"略過(空值) {empty_by_module[module]}｜錯誤 {n_err}")
            continue
        inserted, skipped = upsert_ados(db, collection, docs)
        print(f"[{collection}] 新增 {inserted}｜重複略過 {skipped}｜"
              f"略過(空值) {empty_by_module[module]}｜錯誤 {n_err}")

    export_errors(errors, args.error_output)
    print("\ndone")


if __name__ == "__main__":
    main()
