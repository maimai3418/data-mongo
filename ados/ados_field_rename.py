#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ADOS 欄位驗證 & 重命名腳本
- 讀取 INPUT_DIR 內所有 xlsx，每個檔案含 M1/M2/M3/M4 工作表
- 驗證各工作表是否包含該 module 的完整項目欄位（依 field.json）
- 驗證必要欄位 famid / record_date，含 fallback 組合邏輯
- 重命名欄位為新的 field.json 格式
- 錯誤匯出至 {today}_ados_error.xlsx

用法（從專案根目錄執行）：
  # 預設讀 ados/fields 下的 field.json（依檔名自動對應各 module）
  python ados/ados_field_rename.py <input_dir>

  # 指定其他 field.json 資料夾
  python ados/ados_field_rename.py <input_dir> --field-json-dir <dir>

  # 不覆蓋原檔，另存 {stem}_renamed.xlsx
  python ados/ados_field_rename.py <input_dir> --no-overwrite

  # 覆寫某個 module 的 field.json 檔名
  python ados/ados_field_rename.py <input_dir> --field-json M1=custom_M1.json

  # 其他選項
  python ados/ados_field_rename.py <input_dir> --famid-prefix 4
  python ados/ados_field_rename.py <input_dir> -e errors.xlsx

需求套件：pandas、numpy、openpyxl
"""

import os
import sys
import json
import argparse
import pandas as pd
import numpy as np
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 讓中文/emoji 在 Windows 主控台（cp950）也能正常輸出
try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

# ═══════════════════════════════════════════════════════════════
# CONFIG（預設值；皆可由命令列覆寫）
# ═══════════════════════════════════════════════════════════════
TODAY = date.today().strftime("%Y%m%d")

# 預設 field.json 資料夾：與本腳本同層的 fields/（即 ados/fields）
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_FIELD_JSON_DIR = os.path.join(SCRIPT_DIR, "fields")

# famid 組合公式中的前綴常數
# famid = FAMID_PREFIX * 100000 + site * 10000 + fam * 10 + id
DEFAULT_FAMID_PREFIX = 4

MODULES = ["M1", "M2", "M3", "M4"]

# 欄位名稱 variants（case-insensitive 比對）
SITE_COL = "site"
FAM_COL = "fam"
ID_COL = "id"
FAMILY_COL = "family"

INT_Y_VARIANTS = ["int_yr", "int_y"]
INT_M_COL = "int_m"
INT_D_COL = "int_d"

INTERVIEWER_OLD = "p_interv"

# header 欄位：不參與 module item 驗證的欄位（新名稱）
HEADER_FIELDS = {"famid", "birth_date", "sex", "record_date", "interviewer"}
# ═══════════════════════════════════════════════════════════════


def load_field_specs(field_json_paths):
    """載入各 module 的 field.json（{module: 路徑}），回傳 item rename map 與完整欄位清單。"""
    specs = {}
    for module, path in field_json_paths.items():
        with open(path, "r", encoding="utf-8") as f:
            fields = json.load(f)

        prefix = f"ados_{module.lower()}_"
        item_new_names = [k for k in fields if k.startswith(prefix)]
        # old codebook name = strip prefix, e.g. ados_m1_a1 → a1
        old_to_new = {k[len(prefix):]: k for k in item_new_names}

        specs[module] = {
            "all_new_fields": list(fields.keys()),
            "item_old_to_new": old_to_new,         # {"a1": "ados_m1_a1", ...}
            "expected_items": set(old_to_new.keys()),  # {"a1", "a2", ...}
        }
    return specs


def find_col(columns_lower, candidates):
    """在 columns_lower (list[str]) 中找第一個匹配 candidates 的欄位，回傳原始名稱 index。"""
    if isinstance(candidates, str):
        candidates = [candidates]
    for c in candidates:
        if c in columns_lower:
            return c
    return None


def build_famid_series(df, famid_prefix):
    """嘗試從 family+id 或 site+fam+id 組合 famid，回傳 (Series|None, method_str)。"""
    cols_lower = {c.lower(): c for c in df.columns}

    # 方式 1: family + id
    fam_col = cols_lower.get(FAMILY_COL)
    id_col = cols_lower.get(ID_COL)
    if fam_col and id_col:
        family = pd.to_numeric(df[fam_col], errors="coerce")
        pid = pd.to_numeric(df[id_col], errors="coerce")
        if family.notna().any() and pid.notna().any():
            famid = (family * 10 + pid).astype("Int64").astype(str)
            famid = famid.str.replace("<NA>", "")
            return famid, "family*10+id"

    # 方式 2: site + fam + id
    site_col = cols_lower.get(SITE_COL)
    fam_col2 = cols_lower.get(FAM_COL)
    if site_col and fam_col2 and id_col:
        site = pd.to_numeric(df[cols_lower[SITE_COL]], errors="coerce")
        fam = pd.to_numeric(df[cols_lower[FAM_COL]], errors="coerce")
        pid = pd.to_numeric(df[cols_lower[ID_COL]], errors="coerce")
        if site.notna().any() and fam.notna().any() and pid.notna().any():
            famid = (
                famid_prefix * 100000 + site * 10000 + fam * 10 + pid
            ).astype("Int64").astype(str)
            famid = famid.str.replace("<NA>", "")
            return famid, f"{famid_prefix}*100000+site*10000+fam*10+id"

    return None, None


def normalize_date_col(series):
    """將日期欄位統一轉為 YYYY-MM-DD 文字格式，無法解析者保留原值。"""
    parsed = pd.to_datetime(series, errors="coerce", format="mixed")
    formatted = parsed.dt.strftime("%Y-%m-%d")
    # 無法解析的保留原始字串（或 NaN）
    return formatted.where(parsed.notna(), series)


def build_record_date_series(df):
    """嘗試從 int_y/int_yr + int_m + int_d 組合 record_date，回傳 YYYY-MM-DD 文字。"""
    cols_lower = {c.lower(): c for c in df.columns}

    y_col = None
    for v in INT_Y_VARIANTS:
        if v in cols_lower:
            y_col = cols_lower[v]
            break
    m_col = cols_lower.get(INT_M_COL)
    d_col = cols_lower.get(INT_D_COL)

    if y_col and m_col and d_col:
        y = pd.to_numeric(df[y_col], errors="coerce")
        m = pd.to_numeric(df[m_col], errors="coerce")
        d = pd.to_numeric(df[d_col], errors="coerce")
        dates = pd.to_datetime(
            {"year": y, "month": m, "day": d}, errors="coerce"
        )
        if dates.notna().any():
            text = dates.dt.strftime("%Y-%m-%d").where(dates.notna(), np.nan)
            return text, f"{y_col}+{m_col.lower()}+{d_col.lower()}"

    return None, None


def process_file(filepath, specs, errors, overwrite, famid_prefix):
    """處理單一 xlsx 檔案，回傳是否有任何 sheet 被成功處理。"""
    filename = os.path.basename(filepath)
    try:
        all_sheets = pd.read_excel(filepath, sheet_name=None, dtype=str)
    except Exception as e:
        errors.append({
            "file": filename, "module": "-",
            "error": f"無法讀取檔案: {e}"
        })
        return False

    modified = {}
    any_success = False

    for module in MODULES:
        if module not in all_sheets:
            errors.append({
                "file": filename, "module": module,
                "error": f"工作表 {module} 不存在"
            })
            continue

        df = all_sheets[module]
        spec = specs[module]
        sheet_errors = []
        cols_lower = {c.lower(): c for c in df.columns}

        # ── 1. 驗證 module item 欄位完整性 ──
        expected = spec["expected_items"]
        found_items = expected & set(cols_lower.keys())
        missing_items = expected - found_items
        if missing_items:
            sorted_missing = sorted(missing_items, key=lambda x: (x[0], int(x[1:])))
            sheet_errors.append(
                f"缺少項目欄位: {', '.join(sorted_missing)}"
            )

        # ── 2. 驗證 famid ──
        has_famid = "famid" in cols_lower
        famid_series = None
        famid_method = None
        if not has_famid:
            famid_series, famid_method = build_famid_series(df, famid_prefix)
            if famid_series is None:
                sheet_errors.append(
                    "缺少 famid，且無法從 family+id 或 site+fam+id 組合"
                )

        # ── 3. 驗證 record_date ──
        has_record_date = "record_date" in cols_lower
        rdate_series = None
        rdate_method = None
        if not has_record_date:
            rdate_series, rdate_method = build_record_date_series(df)
            if rdate_series is None:
                sheet_errors.append(
                    "缺少 record_date，且無法從 int_y(r)/int_m/int_d 組合"
                )

        # ── 如果有 blocking error，記錄並跳過此 sheet ──
        if sheet_errors:
            for e in sheet_errors:
                errors.append({"file": filename, "module": module, "error": e})
            continue

        # ── 4. 組合 famid / record_date（如需要）──
        if famid_series is not None:
            df.insert(0, "famid", famid_series)
            print(f"  [{module}] famid 已從 {famid_method} 組合")

        if rdate_series is not None:
            # 插入到 famid 之後
            insert_pos = list(df.columns).index("famid") + 1 if "famid" in df.columns else 1
            df.insert(insert_pos, "record_date", rdate_series)
            print(f"  [{module}] record_date 已從 {rdate_method} 組合")

        # ── 5. 建立 rename map 並執行 ──
        rename_map = {}
        cols_lower_current = {c.lower(): c for c in df.columns}

        # item 欄位: a1 → ados_m1_a1
        for old_lower, new_name in spec["item_old_to_new"].items():
            orig_col = cols_lower_current.get(old_lower)
            if orig_col and orig_col != new_name:
                rename_map[orig_col] = new_name

        # header 欄位
        interv_col = cols_lower_current.get(INTERVIEWER_OLD.lower())
        if interv_col:
            rename_map[interv_col] = "interviewer"

        sex_col = cols_lower_current.get("sex")
        if sex_col and sex_col != "sex":
            rename_map[sex_col] = "sex"

        if rename_map:
            df.rename(columns=rename_map, inplace=True)

        # ── 5b. 日期欄位統一為 YYYY-MM-DD 文字格式 ──
        for date_col in ["record_date", "birth_date"]:
            if date_col in df.columns:
                df[date_col] = normalize_date_col(df[date_col])

        # ── 6. 只保留 field.json 中定義的欄位（加上 famid/record_date）──
        target_fields = spec["all_new_fields"]
        keep_cols = [c for c in df.columns if c in target_fields]
        # 保留順序與 field.json 一致
        ordered = [f for f in target_fields if f in keep_cols]
        df = df[ordered]

        modified[module] = df
        any_success = True

    # ── 寫回檔案 ──
    if any_success:
        if overwrite:
            out_path = filepath
        else:
            stem = Path(filepath).stem
            out_path = os.path.join(os.path.dirname(filepath), f"{stem}_renamed.xlsx")

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            for sheet_name, df in modified.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"  已儲存: {os.path.basename(out_path)} ({', '.join(modified.keys())})")

    return any_success


def export_errors(errors, error_output):
    """匯出錯誤報告至 Excel。"""
    if not errors:
        print(f"\n✓ 無錯誤，不產生 error 檔案")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "errors"

    # Header
    headers = ["file", "module", "error"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="C0392B")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data
    for row_idx, err in enumerate(errors, 2):
        ws.cell(row=row_idx, column=1, value=err["file"])
        ws.cell(row=row_idx, column=2, value=err["module"])
        ws.cell(row=row_idx, column=3, value=err["error"])

    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 60

    wb.save(error_output)
    print(f"\n✗ 共 {len(errors)} 筆錯誤，已匯出: {os.path.basename(error_output)}")


def parse_field_json_overrides(pairs):
    """把 ['M1=foo.json'] 解析成 {module: filename}（只含使用者明確指定者）。"""
    overrides = {}
    for token in pairs or []:
        if "=" not in token:
            raise SystemExit(f"--field-json 格式錯誤：「{token}」，應為 MODULE=檔名（例：M1=foo.json）")
        module, filename = token.split("=", 1)
        module, filename = module.strip().upper(), filename.strip()
        if module not in MODULES:
            raise SystemExit(f"--field-json 模組無效：「{module}」，須為 {', '.join(MODULES)} 其一")
        if not filename:
            raise SystemExit(f"--field-json 檔名不可為空：「{token}」")
        overrides[module] = filename
    return overrides


def discover_field_json(field_json_dir, module):
    """在 field_json_dir 中自動找該 module 的 field.json：
    檔名（小寫）需含 'ados_<module>' 且以 'fields.json' 結尾；多個取檔名最大者（日期最新）。"""
    needle = f"ados_{module.lower()}"
    matches = sorted(
        f for f in os.listdir(field_json_dir)
        if f.lower().endswith("fields.json") and needle in f.lower()
    )
    return matches[-1] if matches else None


def resolve_field_json_files(field_json_dir, overrides):
    """回傳 {module: 絕對路徑}：override 優先，其餘自動偵測；找不到則中止。"""
    paths = {}
    for module in MODULES:
        if module in overrides:
            filename = overrides[module]
        else:
            filename = discover_field_json(field_json_dir, module)
            if filename is None:
                raise SystemExit(
                    f"在 {field_json_dir} 找不到 {module} 的 field.json"
                    f"（檔名需含 ADOS_{module} 且以 fields.json 結尾），"
                    f"或用 --field-json {module}=檔名 指定。"
                )
        paths[module] = os.path.join(field_json_dir, filename)
    return paths


def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="驗證並重命名 ADOS xlsx（M1~M4 工作表）欄位為新的 field.json 格式")
    parser.add_argument("input_dir",
                        help="待處理 xlsx 所在資料夾")
    parser.add_argument("--field-json-dir", default=DEFAULT_FIELD_JSON_DIR, metavar="DIR",
                        help="各 module field.json 所在資料夾（預設 ados/fields）")
    parser.add_argument("--field-json", nargs="+", metavar="MODULE=FILE", default=None,
                        help="覆寫某 module 的 field.json 檔名（例：--field-json M1=foo.json）")
    parser.add_argument("--famid-prefix", type=int, default=DEFAULT_FAMID_PREFIX,
                        metavar="N",
                        help=f"famid 組合前綴常數（預設 {DEFAULT_FAMID_PREFIX}）")
    parser.add_argument("--no-overwrite", action="store_true",
                        help="不覆蓋原檔，改另存 {stem}_renamed.xlsx（預設覆蓋原檔）")
    parser.add_argument("-e", "--error-output", default=None, metavar="XLSX",
                        help="錯誤報告輸出路徑（預設 {input_dir}/{today}_ados_error.xlsx）")
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)

    input_dir = os.path.abspath(args.input_dir)
    if not os.path.isdir(input_dir):
        print(f"Error: 找不到資料夾 {input_dir}")
        sys.exit(1)
    if not os.path.isdir(args.field_json_dir):
        print(f"Error: 找不到 field.json 資料夾 {args.field_json_dir}")
        sys.exit(1)

    overrides = parse_field_json_overrides(args.field_json)
    field_json_paths = resolve_field_json_files(args.field_json_dir, overrides)
    overwrite = not args.no_overwrite
    error_output = args.error_output or os.path.join(input_dir, f"{TODAY}_ados_error.xlsx")

    print("使用 field.json：")
    for module in MODULES:
        print(f"  [{module}] {os.path.basename(field_json_paths[module])}")

    try:
        specs = load_field_specs(field_json_paths)
    except FileNotFoundError as e:
        print(f"Error: 無法載入 field.json：{e}")
        sys.exit(1)

    errors = []
    xlsx_files = sorted([
        f for f in os.listdir(input_dir)
        if f.endswith(".xlsx") and not f.startswith("~$") and not f.startswith(TODAY)
    ])

    if not xlsx_files:
        print(f"INPUT_DIR 中沒有 xlsx 檔案: {input_dir}")
        return

    print(f"找到 {len(xlsx_files)} 個檔案\n")

    for filename in xlsx_files:
        filepath = os.path.join(input_dir, filename)
        print(f"處理: {filename}")
        process_file(filepath, specs, errors, overwrite, args.famid_prefix)

    export_errors(errors, error_output)


if __name__ == "__main__":
    main()
