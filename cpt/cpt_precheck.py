#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CPT 匯入前彙整檢查（precheck）。

彙整多個 CPT xlsx 資料檔，確認「應該是同一筆紀錄」的資料不會被重複匯入資料庫：
  1. 以 famid + age 作為重複判定的 key。
     欄位名稱一律先轉小寫再比對（famid / famID 視為相同）；
     age 欄位另可接受別名 age_CPT / age_cpt。
  2. key 相同時，比對 cpt/fields 中 field.json 內
     除了 sex, birth_date, record_date（與 key 本身 famid）以外的所有欄位：
       - 全部相同 → 視為同一筆，只保留一筆（合併記錄所有資料來源）。
       - 有任何差異 → 視為不同筆，全部保留（各自標記資料來源與差異欄位）。
  3. 輸出 cpt_precheck.xlsx：
       - Summary    檢查彙總（檔案數、總列數、不重複紀錄數、重複移除數…）
       - Files      所有進行 precheck 的檔案清單，標記出現重複資料的檔案
       - Duplicates 完全重複被去重的紀錄與其出現的檔案名稱
       - Conflicts  key 相同但欄位有差異、兩筆皆保留的紀錄
       - Records    彙整後保留的所有資料（含資料來源，可供後續匯入使用）

用法（從專案根目錄執行）：
  python cpt/cpt_precheck.py <檔案或資料夾>
  python cpt/cpt_precheck.py data_dir/
  python cpt/cpt_precheck.py a.xlsx b.xlsx
  python cpt/cpt_precheck.py data_dir/ -o report.xlsx

需求套件：pandas、openpyxl
"""

import os
import sys
import json
import argparse
from collections import OrderedDict

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.utils.wait_and_retry import wait_and_retry

# 讓中文/emoji 在 Windows 主控台（cp950）也能正常輸出
try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

# ── CONFIG ──────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FIELDS_DIR = os.path.join(SCRIPT_DIR, "fields")
DEFAULT_OUTPUT = os.path.join(SCRIPT_DIR, "cpt_precheck.xlsx")

# 重複判定 key（欄位名稱一律先轉小寫再比對，famid / famID 視為相同）
KEY_FIELDS = ["famid", "age"]
# age 欄位在資料檔中可能的別名（小寫），讀檔時統一改名為 age
AGE_ALIASES = ("age_cpt",)
# field.json 中不參與比對的欄位（famid 為 key 本身，另外排除下列欄位）
EXCLUDE_COMPARE = {"famid", "sex", "birth_date", "record_date"}
# ────────────────────────────────────────────────────────────────────────

# ---- 報表樣式（沿用 precheck_upload.py 的風格）----
FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="4472C4")
WARN_FILL = PatternFill("solid", start_color="FFF2CC")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=12)
BASE_FONT = Font(name=FONT_NAME)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WARN_MARK = "⚠️ 有重複"


# ── field.json ──────────────────────────────────────────────────────────
def discover_field_json(fields_dir=FIELDS_DIR):
    """找出 CPT 的 field.json：檔名（小寫）含 'cpt' 且以 'fields.json' 結尾；
    多個取檔名最大者（日期前綴最新）。"""
    if not os.path.isdir(fields_dir):
        return None
    matches = sorted(
        f for f in os.listdir(fields_dir)
        if f.lower().endswith("fields.json") and "cpt" in f.lower()
    )
    return os.path.join(fields_dir, matches[-1]) if matches else None


def load_compare_fields():
    """回傳 (json 路徑, 全欄位清單, 比對欄位清單)，欄位名稱皆轉小寫。"""
    path = discover_field_json()
    if not path:
        raise SystemExit(f"Error: 在 {FIELDS_DIR} 找不到 CPT 的 fields.json")
    with open(path, "r", encoding="utf-8") as f:
        fields = json.load(f)
    # field.json 欄位名稱統一轉小寫（與資料欄位的小寫比對一致）
    json_fields = []
    for k in fields.keys():
        kl = k.strip().lower()
        if kl not in json_fields:
            json_fields.append(kl)
    compare = [c for c in json_fields if c not in EXCLUDE_COMPARE]
    return path, json_fields, compare


# ── 數值正規化 ──────────────────────────────────────────────────────────
def norm_value(v):
    """把儲存格值轉成可比對的形式：空值→None、可轉數字→數字、其餘→去空白字串。

    讓 "5"、"5.0"、5、5.0 視為相同值。
    """
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    s = str(v).strip()
    if s == "":
        return None
    try:
        f = float(s)
    except ValueError:
        return s
    return int(f) if f.is_integer() else f


def norm_key(v):
    """key 欄位轉成字串（空值→''）。"""
    n = norm_value(v)
    return "" if n is None else str(n)


def fmt_sources(sources):
    """來源清單 → 顯示字串，同一來源多次出現以 (xN) 標記。"""
    counts = OrderedDict()
    for s in sources:
        counts[s] = counts.get(s, 0) + 1
    return ", ".join(s if n == 1 else f"{s} (x{n})" for s, n in counts.items())


# ── 報表工具（沿用 precheck_upload.py）────────────────────────────────────
def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autofit(ws, max_width=50):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 0
        for cell in col:
            if cell.value is not None:
                longest = max(longest, max(len(s) for s in str(cell.value).split("\n")))
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), max_width)


def write_table(ws, headers, rows, start_row=1, warn_col=None, warn_value=None):
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h).font = BASE_FONT
    style_header(ws, len(headers), row=start_row)
    for i, rowvals in enumerate(rows, start=start_row + 1):
        flag = warn_col is not None and rowvals[warn_col] == warn_value
        for j, v in enumerate(rowvals, start=1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = BASE_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if flag:
                cell.fill = WARN_FILL
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


# ── 檔案蒐集 ────────────────────────────────────────────────────────────
def collect_files(paths, output_path):
    files = []
    out_ap = os.path.abspath(output_path)
    for p in paths:
        if os.path.isdir(p):
            for f in sorted(os.listdir(p)):
                if f.endswith((".xlsx", ".xls")) and not f.startswith("~$"):
                    files.append(os.path.join(p, f))
        elif os.path.isfile(p):
            files.append(p)
        else:
            print(f"⚠️  略過：找不到 {p}")
    seen, unique = set(), []
    for f in files:
        ap = os.path.abspath(f)
        if ap in seen or ap == out_ap:
            continue
        seen.add(ap)
        unique.append(f)
    return unique


def read_records(files):
    """讀取所有檔案的所有工作表，回傳 (紀錄清單, 欄位出現順序, 讀檔錯誤)。

    每筆紀錄：{"source": 檔名[｜工作表], "row": Excel 列號, "data": {欄位: 原值}}
    """
    records = []
    all_cols = []
    seen_cols = set()
    read_errors = []

    for filepath in files:
        filename = os.path.basename(filepath)
        try:
            all_sheets = pd.read_excel(filepath, sheet_name=None, dtype=str)
        except Exception as e:
            read_errors.append((filename, f"無法讀取檔案: {e}"))
            continue

        multi = len(all_sheets) > 1
        for sheet_name, df in all_sheets.items():
            source = f"{filename}｜{sheet_name}" if multi else filename
            df = df.copy()
            # 欄位名稱統一轉小寫再比對（famid / famID 只是大小寫差異）
            df.columns = df.columns.str.strip().str.lower()
            df = df.loc[:, df.columns.notna()]
            df = df.dropna(how="all")

            # age 欄位可能叫 age_CPT / age_cpt，統一改名為 age
            if "age" not in df.columns:
                alias = next((a for a in AGE_ALIASES if a in df.columns), None)
                if alias:
                    df = df.rename(columns={alias: "age"})

            for col in df.columns:
                if col not in seen_cols:
                    seen_cols.add(col)
                    all_cols.append(col)

            for idx, row in df.iterrows():
                records.append({
                    "source": source,
                    "row": idx + 2,  # +1 標題列、+1 轉 1-based
                    "data": row.to_dict(),
                })

    return records, all_cols, read_errors


# ── 核心比對 ────────────────────────────────────────────────────────────
def dedupe(records, compare_fields):
    """依 famid+age 分組、依比對欄位分辨變異。

    回傳：
      kept        保留的紀錄清單（dict，含 status/sources/rec/diff_fields）
      dup_groups  完全重複群組 [(key, recs)]
      conflicts   差異保留群組 [(key, [(variant_recs, diff_fields)], diff_fields)]
      missing     缺 key 的紀錄清單
    """
    groups = OrderedDict()
    missing = []
    for rec in records:
        key = tuple(norm_key(rec["data"].get(f)) for f in KEY_FIELDS)
        if any(k == "" for k in key):
            missing.append(rec)
            continue
        groups.setdefault(key, []).append(rec)

    kept = []
    dup_groups = []
    conflicts = []

    for key, grp in groups.items():
        variants = OrderedDict()  # signature -> [recs]
        for rec in grp:
            sig = tuple(norm_value(rec["data"].get(f)) for f in compare_fields)
            variants.setdefault(sig, []).append(rec)

        is_conflict = len(variants) > 1

        # 群組內有差異時，找出實際不同的欄位
        diff_fields = []
        if is_conflict:
            sigs = list(variants.keys())
            diff_fields = [
                f for i, f in enumerate(compare_fields)
                if len({sig[i] for sig in sigs}) > 1
            ]

        for sig, recs in variants.items():
            if is_conflict:
                status = "差異保留"
            elif len(recs) > 1:
                status = "重複去重"
            else:
                status = "唯一"
            kept.append({
                "key": key,
                "status": status,
                "rec": recs[0],
                "sources": [f"{r['source']} 列{r['row']}" for r in recs],
                "n_dup": len(recs),
                "diff_fields": diff_fields if is_conflict else [],
            })
            if len(recs) > 1:
                dup_groups.append((key, recs))

        if is_conflict:
            conflicts.append((key, list(variants.items()), diff_fields))

    return kept, dup_groups, conflicts, missing


def source_stats(records, dup_groups, conflicts, missing):
    """統計每個來源（檔案｜工作表）的列數與重複情況。"""
    stats = OrderedDict()

    def entry(source):
        return stats.setdefault(
            source, {"rows": 0, "missing": 0, "dup": 0, "conflict": 0})

    for rec in records:
        entry(rec["source"])["rows"] += 1
    for rec in missing:
        entry(rec["source"])["missing"] += 1
    for _, recs in dup_groups:
        for rec in recs:
            entry(rec["source"])["dup"] += 1
    for _, variants, _ in conflicts:
        for _, recs in variants:
            for rec in recs:
                entry(rec["source"])["conflict"] += 1
    return stats


# ── 輸出 ────────────────────────────────────────────────────────────────
def build_report(output_path, files, json_path, json_fields, compare_fields,
                 records, all_cols, read_errors, kept, dup_groups, conflicts,
                 missing):
    n_removed = sum(len(recs) - 1 for _, recs in dup_groups)
    n_conflict_rows = sum(len(v) for _, v, _ in conflicts)
    n_unique = len(kept)

    wb = Workbook()

    # ---- Summary ----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "CPT 匯入前彙整檢查"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (f"重複判定 key：{' + '.join(KEY_FIELDS)}｜"
                f"比對欄位：{len(compare_fields)} 個"
                f"（{os.path.basename(json_path)}，"
                f"排除 {', '.join(sorted(EXCLUDE_COMPARE - {'famid'}))}）")
    ws["A2"].font = BASE_FONT
    summary_rows = [
        ("檢查檔案數", len(files)),
        ("讀取失敗檔案數", len(read_errors)),
        ("總列數", len(records)),
        ("缺 key（famid/age）列數", len(missing)),
        ("彙整後不重複紀錄數", n_unique),
        ("完全重複移除筆數", n_removed),
        ("完全重複群組數", len(dup_groups)),
        ("欄位差異保留群組數", len(conflicts)),
        ("欄位差異保留筆數", n_conflict_rows),
    ]
    write_table(ws, ["項目", "數值"], summary_rows, start_row=4)
    autofit(ws)

    # ---- Files ----
    ws = wb.create_sheet("Files")
    stats = source_stats(records, dup_groups, conflicts, missing)
    file_rows = []
    for source, s in stats.items():
        mark = WARN_MARK if (s["dup"] or s["conflict"]) else ""
        file_rows.append((source, s["rows"], s["missing"],
                          s["dup"], s["conflict"], mark))
    for filename, err in read_errors:
        file_rows.append((filename, 0, 0, 0, 0, f"讀取失敗：{err}"))
    if not file_rows:
        file_rows = [("（無）", 0, 0, 0, 0, "")]
    write_table(
        ws,
        ["檔案（｜工作表）", "列數", "缺key列數", "完全重複列數", "差異保留列數", "標記"],
        file_rows, warn_col=5, warn_value=WARN_MARK,
    )
    autofit(ws, max_width=60)

    # ---- Duplicates ----
    ws = wb.create_sheet("Duplicates")
    dup_rows = []
    for key, recs in dup_groups:
        sources = [f"{r['source']} 列{r['row']}" for r in recs]
        dup_rows.append((*key, len(recs), sources[0], fmt_sources(sources)))
    if not dup_rows:
        dup_rows = [("（無）", "", "", "", "沒有完全重複的紀錄")]
    write_table(
        ws,
        [*KEY_FIELDS, "重複筆數", "保留來源", "重複出現檔案"],
        dup_rows,
    )
    autofit(ws, max_width=70)

    # ---- Conflicts ----
    ws = wb.create_sheet("Conflicts")
    conflict_rows = []
    for key, variants, diff_fields in conflicts:
        for vi, (_sig, recs) in enumerate(variants, start=1):
            detail = "; ".join(
                f"{f}={norm_value(recs[0]['data'].get(f))}" for f in diff_fields)
            sources = [f"{r['source']} 列{r['row']}" for r in recs]
            conflict_rows.append((
                *key, f"{vi}/{len(variants)}", fmt_sources(sources),
                ", ".join(diff_fields), detail,
            ))
    if not conflict_rows:
        conflict_rows = [("（無）", "", "", "", "", "沒有 key 相同但欄位有差異的紀錄")]
    write_table(
        ws,
        [*KEY_FIELDS, "變異", "資料來源", "差異欄位", "差異內容"],
        conflict_rows,
    )
    autofit(ws, max_width=70)

    # ---- Records（彙整後保留的資料）----
    ws = wb.create_sheet("Records")
    present = set(all_cols)
    # 欄位順序：famid、age 之後照 field.json 順序，再接其餘欄位
    data_cols = [c for c in json_fields if c in present and c != "famid"]
    data_cols += [c for c in all_cols
                  if c not in json_fields and c not in ("famid", "age")]
    headers = ["狀態", "資料來源", "famid", "age", *data_cols]

    def record_row(status, sources, rec):
        d = rec["data"]
        return (status, fmt_sources(sources),
                norm_key(d.get("famid")), norm_key(d.get("age")),
                *[norm_value(d.get(c)) for c in data_cols])

    record_rows = [record_row(k["status"], k["sources"], k["rec"]) for k in kept]
    record_rows += [record_row("缺key", [f"{r['source']} 列{r['row']}"], r)
                    for r in missing]
    if not record_rows:
        record_rows = [("（無）", *[""] * (len(headers) - 1))]
    write_table(ws, headers, record_rows, warn_col=0, warn_value="差異保留")
    autofit(ws, max_width=40)

    wait_and_retry(lambda: wb.save(output_path), output_path)
    return summary_rows


def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="CPT 匯入前彙整檢查：依 famid+age 找出重複資料並彙整")
    parser.add_argument("paths", nargs="+",
                        help="要檢查的 xlsx 檔案或資料夾")
    parser.add_argument("-o", "--output", default=DEFAULT_OUTPUT, metavar="XLSX",
                        help=f"報表輸出路徑（預設 {DEFAULT_OUTPUT}）")
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)

    files = collect_files(args.paths, args.output)
    if not files:
        print("Error: 找不到任何 xlsx 檔案")
        sys.exit(1)

    json_path, json_fields, compare_fields = load_compare_fields()
    print(f"📄 檔案數：{len(files)}")
    print(f"📐 比對欄位來源：{json_path}")
    print(f"   重複判定 key：{' + '.join(KEY_FIELDS)}｜比對欄位：{len(compare_fields)} 個"
          f"（排除 {', '.join(sorted(EXCLUDE_COMPARE - {'famid'}))}）")

    records, all_cols, read_errors = read_records(files)
    for filename, err in read_errors:
        print(f"⚠️  {filename}: {err}")
    print(f"共讀取 {len(records)} 列")

    kept, dup_groups, conflicts, missing = dedupe(records, compare_fields)

    summary_rows = build_report(
        args.output, files, json_path, json_fields, compare_fields,
        records, all_cols, read_errors, kept, dup_groups, conflicts, missing)

    # ---- 主控台摘要 ----
    print("\n===== 檢查結果 =====")
    for label, value in summary_rows:
        print(f"{label}：{value}")
    if dup_groups:
        print("\n⚠️ 完全重複（只保留一筆）：")
        for key, recs in dup_groups:
            sources = fmt_sources(f"{r['source']} 列{r['row']}" for r in recs)
            print(f"  famid={key[0]}, age={key[1]}（{len(recs)} 筆）：{sources}")
    if conflicts:
        print("\n⚠️ key 相同但欄位有差異（全部保留）：")
        for key, variants, diff_fields in conflicts:
            print(f"  famid={key[0]}, age={key[1]}（{len(variants)} 種變異）："
                  f"差異欄位 {', '.join(diff_fields)}")
    print(f"\n已輸出報表：{args.output}")


if __name__ == "__main__":
    main()
