#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CPT 匯入：讀取 CPT xlsx，依 famid + record_date 做三層去重複後匯入 CPT collection。

三層去重複（CLAUDE.md §5）：
  1. hard dedup      同一次執行內，famid+record_date 相同且所有量表欄位也相同的列
                      → 只保留一筆。
  2. retest 保留     famid 相同但 record_date 在 DB 中不存在（不同施測日期）
                      → 視為重測，直接新增，不與其他日期的紀錄比較。
  3. outer join      famid+record_date 與 DB 現有紀錄相同、但量表欄位值不同
     人工審查         → 無法自動判定何者為真，兩邊皆保留並匯出衝突報告。

沒有 record_date 的列（CLAUDE.md §3 no-date 慣例）：
  - 不強制配對日期，也不套用上述三層 key 比對邏輯（record_date 為 null，
    無法可靠比對是否為同一次施測）。
  - 以 famid + 量表欄位值 是否與既有 null-date 紀錄完全相同 做 hard dedup；
    完全相同 → 略過，其餘一律新增（record_date=null, flag_no_date=true）。
  - 所有處理過的 no-date 列（新增或略過）皆記錄到 no_date_records.xlsx，
    供之後 famid_timepoint_map.py（日期回填管線，尚未實作）人工/自動回填參考。

CPT 沒有 999 null sentinel 慣例，不做全域 999→null 轉換（見 cpt_config.py）。
famid 在 CPT 語境直接等同個人 ID（CLAUDE.md §2）。

用法（從專案根目錄執行）：
  python cpt/cpt_import.py <檔案或資料夾>
  python cpt/cpt_import.py data_dir/
  python cpt/cpt_import.py a.xlsx b.xlsx
  python cpt/cpt_import.py data_dir/ --dry-run
  python cpt/cpt_import.py a.xlsx --project-code NHRI113
  python cpt/cpt_import.py a.xlsx -e errors.xlsx --conflict-output conflicts.xlsx

需求套件：pandas、openpyxl、pymongo、python-dotenv
"""

import os
import sys
import argparse
from collections import OrderedDict
from datetime import date

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.importer import get_db
from src.no_date_writer import write_no_date_xlsx
from cpt_config import (
    CPT_COLLECTION,
    CPT_SHARED_FIELDS,
    CPT_HEADER_FIELDS,
    CPT_FIELDS_DIR,
    load_field_rules,
)

try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

load_dotenv()

TODAY = date.today().strftime("%Y%m%d")
DEFAULT_ERROR_OUTPUT = f"{TODAY}_CPT_import_error.xlsx"
DEFAULT_CONFLICT_OUTPUT = f"{TODAY}_CPT_conflict_report.xlsx"
PROJECT_FIELD = "research_project_code"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="4472C4")
WARN_FILL = PatternFill("solid", start_color="FFF2CC")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
BASE_FONT = Font(name=FONT_NAME)


# ── 數值處理 ────────────────────────────────────────────────────────────
def is_blank(val):
    return val is None or (isinstance(val, float) and pd.isna(val))


def to_number(s):
    try:
        f = float(s)
    except (ValueError, TypeError):
        return None
    return int(f) if f.is_integer() else f


def clean_value(val):
    if is_blank(val):
        return None
    s = str(val).strip()
    return s if s != "" else None


def to_date_text(col, val):
    if is_blank(val):
        return None, None
    s = str(val).strip()
    if s == "":
        return None, None
    dt = pd.to_datetime(s, format="mixed", dayfirst=False, errors="coerce")
    if pd.isna(dt):
        return s, f"{col} 日期格式錯誤"
    return dt.strftime("%Y-%m-%d"), None


def convert_item(col, val, rule):
    """量表欄位轉換，回傳 (converted, error)。空值→null；有 rule 做型別+範圍驗證；
    無 rule 原樣上傳（能轉數字就存數字）。不做 999→null 轉換（CPT 無此慣例）。"""
    if is_blank(val):
        return None, None
    s = str(val).strip()
    if s == "":
        return None, None

    num = to_number(s)

    if rule is None:
        return (num if num is not None else s), None

    rtype = rule.get("type")
    if rtype in ("int", "float"):
        if num is None:
            return None, f"{col} 數值錯誤"
        if rtype == "int" and not isinstance(num, int):
            return None, f"{col} 數值錯誤(非整數)"
        if "range" in rule:
            lo, hi = rule["range"]
            if not (lo <= num <= hi):
                return None, f"{col} 數值錯誤({num} 不在 {lo}–{hi})"
        return num, None

    return (num if num is not None else s), None


def norm_value(v):
    """把值轉成可比對的形式：空值→None、可轉數字→數字、其餘→去空白字串。
    讓 "5"、"5.0"、5、5.0 視為相同值（同 cpt_precheck.py）。"""
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    s = str(v).strip()
    if s == "":
        return None
    try:
        f = float(s)
    except (ValueError, TypeError):
        return s
    return int(f) if f.is_integer() else f


def famid_str(val):
    """famid 永遠字串化，清除 Excel 浮點數殘留（CLAUDE.md §2）。"""
    if is_blank(val):
        return None
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s if s != "" else None


# ── 單一工作表處理 ──────────────────────────────────────────────────────
def process_sheet(filename, sheet_name, df, errors, field_rules):
    """處理單一工作表。

    回傳 (dated_records, no_date_records, n_empty)：
      dated_records / no_date_records：[{"source":..., "doc": {...}}]
    """
    df = df.copy()
    df.columns = df.columns.str.strip()
    df = df.loc[:, df.columns.notna()]
    df = df.dropna(how="all")

    header_set = set(CPT_HEADER_FIELDS)
    item_cols = [c for c in df.columns if c not in header_set]
    header_cols = [c for c in df.columns if c in header_set]
    source = f"{filename}｜{sheet_name}"

    dated_records = []
    no_date_records = []
    n_empty = 0

    for _, row in df.iterrows():
        rowd = row.to_dict()
        doc = {}
        row_errors = []

        for col in header_cols:
            if col == "famid":
                doc["famid"] = famid_str(rowd.get("famid"))
                continue
            rule = field_rules.get(col)
            if rule and rule.get("type") == "date":
                new_val, err = to_date_text(col, rowd.get(col))
                if err:
                    row_errors.append(err)
                doc[col] = new_val
            else:
                new_val, err = convert_item(col, rowd.get(col), rule)
                if err:
                    row_errors.append(err)
                else:
                    doc[col] = new_val

        for col in item_cols:
            rule = field_rules.get(col)
            new_val, err = convert_item(col, rowd.get(col), rule)
            if err:
                row_errors.append(err)
            else:
                doc[col] = new_val

        if item_cols and all(doc.get(c) is None for c in item_cols):
            n_empty += 1
            continue

        if doc.get("famid") is None:
            row_errors.append("缺少 famid")

        if row_errors:
            errors.append({
                "file": filename,
                "sheet": sheet_name,
                "famid": doc.get("famid"),
                "record_date": doc.get("record_date"),
                "error": "; ".join(row_errors),
            })
            continue

        record = {"source": source, "doc": doc}
        if doc.get("record_date"):
            dated_records.append(record)
        else:
            no_date_records.append(record)

    return dated_records, no_date_records, n_empty


def build_signature(doc, compare_fields):
    return tuple(norm_value(doc.get(f)) for f in compare_fields)


def fmt_sources(sources):
    counts = OrderedDict()
    for s in sources:
        counts[s] = counts.get(s, 0) + 1
    return ", ".join(s if n == 1 else f"{s} (x{n})" for s, n in counts.items())


# ── 三層去重複 + DB 比對 ─────────────────────────────────────────────────
def reconcile(dated_records, no_date_records, compare_fields, db):
    """回傳 dict：
      to_insert          [doc, ...]                 待新增（含 no-date）
      exact_dup_count    int                         與 DB 完全相同、略過
      batch_conflicts    [{...}]                     同批次內 key 相同但值不同
      value_conflicts    [{...}]                     與 DB 現有紀錄 key 相同但值不同
      no_date_report     [{...}]                     no-date 列處理結果（供 no_date_records.xlsx）
    """
    # ── Layer 1：批次內 hard dedup（有日期）──
    groups = OrderedDict()
    for rec in dated_records:
        key = tuple(rec["doc"][f] for f in CPT_SHARED_FIELDS)
        groups.setdefault(key, []).append(rec)

    kept = {}          # key -> (doc, sources)
    batch_conflicts = []
    for key, recs in groups.items():
        variants = OrderedDict()
        for rec in recs:
            sig = build_signature(rec["doc"], compare_fields)
            variants.setdefault(sig, []).append(rec)
        if len(variants) > 1:
            batch_conflicts.append((key, variants))
            continue
        sig, recs_same = next(iter(variants.items()))
        kept[key] = (recs_same[0]["doc"], [r["source"] for r in recs_same])

    # ── Layer 2/3：查 DB 現有紀錄（一次查詢涵蓋所有 famid）──
    all_famids = sorted({k[0] for k in kept} | {r["doc"]["famid"] for r in no_date_records})
    existing_dated = {}      # (famid, record_date) -> doc
    existing_nodate_sigs = {}  # famid -> set(signature)
    if all_famids:
        for existing in db[CPT_COLLECTION].find({"famid": {"$in": all_famids}}):
            fid = existing.get("famid")
            rdate = existing.get("record_date")
            if rdate:
                existing_dated.setdefault((fid, rdate), existing)
            else:
                existing_nodate_sigs.setdefault(fid, set()).add(
                    build_signature(existing, compare_fields))

    to_insert = []
    value_conflicts = []
    exact_dup_count = 0

    for key, (doc, sources) in kept.items():
        existing = existing_dated.get(key)
        if existing is None:
            to_insert.append(doc)
        elif build_signature(existing, compare_fields) == build_signature(doc, compare_fields):
            exact_dup_count += 1
        else:
            value_conflicts.append({
                "famid": key[0], "record_date": key[1],
                "sources": fmt_sources(sources),
                "db_doc": existing, "new_doc": doc,
            })

    # ── no-date：famid + 量表欄位簽章 hard dedup（不比對日期）──
    no_date_report = []
    seen_this_run = {}  # famid -> set(signature)，避免同批次內重複新增
    for rec in no_date_records:
        doc = rec["doc"]
        fid = doc["famid"]
        sig = build_signature(doc, compare_fields)
        known = existing_nodate_sigs.get(fid, set()) | seen_this_run.get(fid, set())
        if sig in known:
            no_date_report.append({
                "collection": CPT_COLLECTION,
                "reason": "無 record_date，且量表欄位與既有 null-date 紀錄完全相同，略過",
                **doc,
            })
            continue
        doc["flag_no_date"] = True
        to_insert.append(doc)
        seen_this_run.setdefault(fid, set()).add(sig)
        no_date_report.append({
            "collection": CPT_COLLECTION,
            "reason": "無 record_date，已匯入並標記 flag_no_date，"
                      "待後續日期回填管線（famid_timepoint_map.py，尚未實作）處理",
            **doc,
        })

    return {
        "to_insert": to_insert,
        "exact_dup_count": exact_dup_count,
        "batch_conflicts": batch_conflicts,
        "value_conflicts": value_conflicts,
        "no_date_report": no_date_report,
    }


# ── 匯出 ────────────────────────────────────────────────────────────────
def export_errors(errors, output_path):
    if not errors:
        print("\n✓ 無錯誤，不產生 error 檔案")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "errors"
    headers = ["file", "sheet", "famid", "record_date", "error"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row_idx, err in enumerate(errors, 2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=err.get(key))
    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"\n✗ 共 {len(errors)} 筆錯誤，已匯出: {os.path.basename(output_path)}")


def export_conflicts(batch_conflicts, value_conflicts, compare_fields, output_path):
    if not batch_conflicts and not value_conflicts:
        print("✓ 無衝突記錄，不產生 conflict 檔案")
        return

    wb = Workbook()

    # ---- 同批次內衝突 ----
    ws = wb.active
    ws.title = "BatchConflicts"
    headers = ["famid", "record_date", "變異數", "資料來源", "差異欄位", "差異內容"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row_i = 2
    for key, variants in batch_conflicts:
        sigs = list(variants.keys())
        diff_fields = [
            f for i, f in enumerate(compare_fields)
            if len({sig[i] for sig in sigs}) > 1
        ]
        for vi, (sig, recs) in enumerate(variants.items(), start=1):
            detail = "; ".join(
                f"{f}={norm_value(recs[0]['doc'].get(f))}" for f in diff_fields)
            sources = fmt_sources([r["source"] for r in recs])
            ws.cell(row=row_i, column=1, value=key[0]).font = BASE_FONT
            ws.cell(row=row_i, column=2, value=key[1]).font = BASE_FONT
            ws.cell(row=row_i, column=3, value=f"{vi}/{len(variants)}").font = BASE_FONT
            ws.cell(row=row_i, column=4, value=sources).font = BASE_FONT
            ws.cell(row=row_i, column=5, value=", ".join(diff_fields)).font = BASE_FONT
            ws.cell(row=row_i, column=6, value=detail).font = BASE_FONT
            for c in range(1, 7):
                ws.cell(row=row_i, column=c).fill = WARN_FILL
            row_i += 1
    ws.freeze_panes = "A2"
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = 30

    # ---- 與 DB 現有紀錄衝突（value_conflict）----
    ws2 = wb.create_sheet("ValueConflicts")
    headers2 = ["famid", "record_date", "資料來源", "差異欄位", "DB 現有值", "本次新值"]
    for j, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=j, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row_i = 2
    for c in value_conflicts:
        db_doc, new_doc = c["db_doc"], c["new_doc"]
        diff_fields = [
            f for f in compare_fields
            if norm_value(db_doc.get(f)) != norm_value(new_doc.get(f))
        ]
        db_vals = "; ".join(f"{f}={norm_value(db_doc.get(f))}" for f in diff_fields)
        new_vals = "; ".join(f"{f}={norm_value(new_doc.get(f))}" for f in diff_fields)
        ws2.cell(row=row_i, column=1, value=c["famid"]).font = BASE_FONT
        ws2.cell(row=row_i, column=2, value=c["record_date"]).font = BASE_FONT
        ws2.cell(row=row_i, column=3, value=c["sources"]).font = BASE_FONT
        ws2.cell(row=row_i, column=4, value=", ".join(diff_fields)).font = BASE_FONT
        ws2.cell(row=row_i, column=5, value=db_vals).font = BASE_FONT
        ws2.cell(row=row_i, column=6, value=new_vals).font = BASE_FONT
        for c_idx in range(1, 7):
            ws2.cell(row=row_i, column=c_idx).fill = WARN_FILL
        row_i += 1
    ws2.freeze_panes = "A2"
    for col in ws2.columns:
        letter = get_column_letter(col[0].column)
        ws2.column_dimensions[letter].width = 30

    wb.save(output_path)
    n_total = sum(len(v) for _, v in batch_conflicts) + len(value_conflicts)
    print(f"⚠️  共 {n_total} 筆衝突（待人工審查），已匯出: {os.path.basename(output_path)}")


# ── 檔案蒐集 ────────────────────────────────────────────────────────────
def collect_files(paths):
    files = []
    for p in paths:
        if os.path.isdir(p):
            for f in sorted(os.listdir(p)):
                if (f.endswith((".xlsx", ".xls")) and not f.startswith("~$")
                        and not f.startswith(TODAY)):
                    files.append(os.path.join(p, f))
        elif os.path.isfile(p):
            files.append(p)
        else:
            print(f"⚠️  略過：找不到 {p}")
    seen, unique = set(), []
    for f in files:
        ap = os.path.abspath(f)
        if ap not in seen:
            seen.add(ap)
            unique.append(f)
    return unique


def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="CPT 匯入：讀取 xlsx，三層去重複後匯入 CPT collection")
    parser.add_argument("paths", nargs="+", help="要匯入的 xlsx 檔案或資料夾")
    parser.add_argument("--project-code", default=None, metavar="CODE",
                        help="替每筆資料標註計畫代碼（寫入 research_project_code）")
    parser.add_argument("--dry-run", action="store_true",
                        help="只驗證、去重複、比對 DB，不寫入 MongoDB")
    parser.add_argument("-e", "--error-output", default=DEFAULT_ERROR_OUTPUT,
                        metavar="XLSX", help=f"錯誤報告輸出路徑（預設 {DEFAULT_ERROR_OUTPUT}）")
    parser.add_argument("--conflict-output", default=DEFAULT_CONFLICT_OUTPUT,
                        metavar="XLSX", help=f"衝突報告輸出路徑（預設 {DEFAULT_CONFLICT_OUTPUT}）")
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)

    files = collect_files(args.paths)
    if not files:
        print("Error: 找不到任何 xlsx 檔案")
        sys.exit(1)

    project_code = (args.project_code or "").strip() or None
    field_rules = load_field_rules()
    compare_fields = [c for c in field_rules if c not in CPT_HEADER_FIELDS]

    print(f"📄 檔案數：{len(files)}｜collection：{CPT_COLLECTION}"
          f"{'｜DRY-RUN（不寫入）' if args.dry_run else ''}")
    print(f"📐 欄位限制來源：{CPT_FIELDS_DIR}｜比對欄位：{len(compare_fields)} 個")
    if project_code:
        print(f"🏷️  計畫代碼：{project_code}")

    errors = []
    all_dated, all_no_date = [], []
    total_empty = 0

    for filepath in files:
        filename = os.path.basename(filepath)
        print(f"\n處理: {filename}")
        try:
            all_sheets = pd.read_excel(filepath, sheet_name=None, dtype=str)
        except Exception as e:
            errors.append({"file": filename, "sheet": "-", "famid": "",
                           "record_date": "", "error": f"無法讀取檔案: {e}"})
            continue

        for sheet_name, df in all_sheets.items():
            dated, no_date, n_empty = process_sheet(
                filename, sheet_name, df, errors, field_rules)
            all_dated.extend(dated)
            all_no_date.extend(no_date)
            total_empty += n_empty
            print(f"  [{sheet_name}] 有日期 {len(dated)}｜無日期 {len(no_date)}｜"
                  f"略過(空值) {n_empty}")

    db = get_db()
    result = reconcile(all_dated, all_no_date, compare_fields, db)

    to_insert = result["to_insert"]
    if project_code:
        for doc in to_insert:
            doc[PROJECT_FIELD] = project_code

    print("\n===== 匯入結果 =====")
    if args.dry_run:
        print(f"[{CPT_COLLECTION}] dry-run：待新增 {len(to_insert)}｜"
              f"重複略過 {result['exact_dup_count']}｜"
              f"批次內衝突 {sum(len(v) for _, v in result['batch_conflicts'])}｜"
              f"與DB衝突 {len(result['value_conflicts'])}｜"
              f"略過(空值) {total_empty}｜錯誤 {len(errors)}")
    else:
        inserted = 0
        if to_insert:
            db[CPT_COLLECTION].insert_many(to_insert)
            inserted = len(to_insert)
        print(f"[{CPT_COLLECTION}] 新增 {inserted}｜"
              f"重複略過 {result['exact_dup_count']}｜"
              f"批次內衝突 {sum(len(v) for _, v in result['batch_conflicts'])}｜"
              f"與DB衝突 {len(result['value_conflicts'])}｜"
              f"略過(空值) {total_empty}｜錯誤 {len(errors)}")

    export_errors(errors, args.error_output)
    export_conflicts(result["batch_conflicts"], result["value_conflicts"],
                     compare_fields, args.conflict_output)
    if result["no_date_report"]:
        write_no_date_xlsx(result["no_date_report"])

    print("\ndone")


if __name__ == "__main__":
    main()
