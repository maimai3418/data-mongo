"""COLLECTION_MAP 各 collection 的 composite key 重複檢查（read-only）

以 config.SHARED_FIELDS（role + famid + record_date）為 key，
掃描 DB 中 COLLECTION_MAP 全部 collection，找出 key 重複的紀錄。
只報告不修改；有重複時輸出 {TODAY}_dup_key_report.xlsx 供人工判定。

用法：
    python tools/check_duplicate_keys.py
"""
import os
import sys

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from config import COLLECTION_MAP, SHARED_FIELDS
from src.importer import get_db
from src.utils.wait_and_retry import wait_and_retry

# === CONFIG ===
load_dotenv()

OUTPUT_DIR = Path("./output/")
TODAY = datetime.now().strftime("%Y%m%d")
# === END CONFIG ===

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BOLD = Font(bold=True)

REPORT_COLS = ["collection"] + SHARED_FIELDS + ["count", "flag_no_date", "ids"]


def find_duplicates(col):
    """回傳 key 重複的 group 清單：[{_id: {key...}, count, ids}, ...]"""
    pipeline = [
        {"$group": {
            "_id": {k: f"${k}" for k in SHARED_FIELDS},
            "count": {"$sum": 1},
            "ids": {"$push": "$_id"},
        }},
        {"$match": {"count": {"$gt": 1}}},
    ]
    return list(col.aggregate(pipeline, allowDiskUse=True))


def write_report(rows, path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "dup_keys"
    ws.append(REPORT_COLS)
    for cell in ws[1]:
        cell.fill = RED_FILL
        cell.font = BOLD
    for row in rows:
        ws.append([row.get(c) for c in REPORT_COLS])
    ws.freeze_panes = "A2"
    wait_and_retry(lambda: wb.save(path), str(path))


def main():
    db = get_db()
    existing = set(db.list_collection_names())

    dup_rows = []
    dirty = []      # (collection, 重複 key 組數, 涉及筆數)
    not_in_db = []

    for name in COLLECTION_MAP:
        if name not in existing:
            not_in_db.append(name)
            print(f"[{name}] not_in_db")
            continue

        col = db[name]
        total = col.count_documents({})
        groups = find_duplicates(col)
        dup_docs = sum(g["count"] for g in groups)
        print(f"[{name}] 總筆數: {total}, 重複 key 組數: {len(groups)}, 涉及筆數: {dup_docs}")

        if not groups:
            continue
        dirty.append((name, len(groups), dup_docs))
        for g in groups:
            key = g["_id"]
            dup_rows.append({
                "collection": name,
                **{k: key.get(k) for k in SHARED_FIELDS},
                "count": g["count"],
                "flag_no_date": key.get("record_date") in (None, ""),
                "ids": ", ".join(str(i) for i in g["ids"]),
            })

    print("\n=== 總結 ===")
    checked = len(COLLECTION_MAP) - len(not_in_db)
    print(f"檢查 {checked}/{len(COLLECTION_MAP)} 個 collection")
    if not_in_db:
        print(f"不存在於 DB（{len(not_in_db)}）: {', '.join(not_in_db)}")
    if dirty:
        print(f"發現重複（{len(dirty)} 個 collection）:")
        for name, n_groups, n_docs in dirty:
            print(f"  [{name}] 重複 key 組數: {n_groups}, 涉及筆數: {n_docs}")
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        report_path = OUTPUT_DIR / f"{TODAY}_dup_key_report.xlsx"
        write_report(dup_rows, report_path)
        print(f"報告已輸出: {report_path}")
    else:
        print("PASS：所有 collection 皆無 composite key 重複")


if __name__ == "__main__":
    main()
