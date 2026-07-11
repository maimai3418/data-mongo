"""nsc96 系列 research_project_code 改名（加上 asd）— 一次性遷移腳本

用法：
    python tools/rename_nsc96_project_code.py            # dry-run：只統計，輸出 precheck Excel
    python tools/rename_nsc96_project_code.py --apply    # 實際更新（先跑過 dry-run 並人工確認）
"""
import os
import sys

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

import argparse
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.importer import get_db
from src.utils.wait_and_retry import wait_and_retry

# === CONFIG ===
load_dotenv()

OUTPUT_DIR = Path("./output/")
TODAY = datetime.now().strftime("%Y%m%d")

COLLECTIONS = """AAQOL APGAR AQ ARI ASRI ASRS BRIEF BRIEF-A BRIEF-S C-SBEQ CAST CBCL CEQ
CES-D CPRS CTRS DAS DCGII DCGIS DPBI EPQ ERQ-A ERQ-CA ESQ FACES FFQ GSQ MPBI MPI
PCGII PCGIS PMI PSS RAADS-R RBS-R SAICA SCQ SDQ SF-36 SNAP4 SRS SSP SUB TAS-20
TPQ WFIRS-P WFIRS-S WHOQOL YSR""".split()

RENAME_MAP = {
    "nsc96_prc": "nsc96_asd_prc",
    "nsc96_crs": "nsc96_asd_crs",
    "nsc96_trc": "nsc96_asd_trc",
    "nsc96_prs_mom": "nsc96_asd_prs_mom",
    "nsc96_prs_dad": "nsc96_asd_prs_dad",
    "nsc96fu_prc": "nsc96_asd_fu_prc",
    "nsc96fu_crs": "nsc96_asd_fu_crs",
}
# === END CONFIG ===


def scan(db):
    """統計各 collection 內 RENAME_MAP 舊值的筆數，並檢查地圖外變形值與既存新值。"""
    rows = []          # (collection, old, new, count)
    unexpected = []    # (collection, value, count) 不在 RENAME_MAP 的 nsc96 變形
    new_exists = []    # (collection, value, count) 新值已存在
    known = set(RENAME_MAP) | set(RENAME_MAP.values())

    for name in COLLECTIONS:
        col = db[name]
        for old, new in RENAME_MAP.items():
            n = col.count_documents({"research_project_code": old})
            if n:
                rows.append((name, old, new, n))
            n_new = col.count_documents({"research_project_code": new})
            if n_new:
                new_exists.append((name, new, n_new))
        for value in col.distinct("research_project_code",
                                  {"research_project_code": {"$regex": "nsc96", "$options": "i"}}):
            if value not in known:
                unexpected.append((name, value,
                                   col.count_documents({"research_project_code": value})))
    return rows, unexpected, new_exists


def print_report(rows, unexpected, new_exists):
    print(f"{'collection':<10} {'old':<16} {'new':<20} {'count':>6}")
    print("-" * 56)
    for name, old, new, n in rows:
        print(f"{name:<10} {old:<16} {new:<20} {n:>6}")
    print("-" * 56)
    print(f"合計 {sum(r[3] for r in rows)} 筆，{len({r[0] for r in rows})} 個 collection")

    if unexpected:
        print("\n[WARN] 發現不在 RENAME_MAP 的 nsc96 變形值（不會被改名，請人工確認）：")
        for name, value, n in unexpected:
            print(f"  {name}: {value} ({n} 筆)")
    if new_exists:
        print("\n[WARN] 新值已存在於 DB（改名後將與這些記錄混在一起，請先確認來源）：")
        for name, value, n in new_exists:
            print(f"  {name}: {value} ({n} 筆)")


def write_precheck_xlsx(rows, unexpected, new_exists):
    OUTPUT_DIR.mkdir(exist_ok=True)
    out_path = OUTPUT_DIR / f"{TODAY}_nsc96_rename_precheck.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "rename_map"
    header_fill = PatternFill("solid", fgColor="4472C4")
    warn_fill = PatternFill("solid", fgColor="C00000")
    header_font = Font(bold=True, color="FFFFFF")

    ws.append(["collection", "old_value", "new_value", "count"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append(list(r))
    ws.append(["TOTAL", "", "", sum(r[3] for r in rows)])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for col_letter, w in zip("ABCD", (12, 18, 22, 8)):
        ws.column_dimensions[col_letter].width = w
    ws.freeze_panes = "A2"

    if unexpected or new_exists:
        ws2 = wb.create_sheet("warnings")
        ws2.append(["type", "collection", "value", "count"])
        for cell in ws2[1]:
            cell.fill = warn_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for name, value, n in unexpected:
            ws2.append(["不在對照表的 nsc96 變形值", name, value, n])
        for name, value, n in new_exists:
            ws2.append(["新值已存在", name, value, n])
        for col_letter, w in zip("ABCD", (26, 12, 22, 8)):
            ws2.column_dimensions[col_letter].width = w

    wait_and_retry(lambda: wb.save(out_path), str(out_path))
    print(f"\n已輸出 precheck: {out_path}")


def apply_rename(db, rows):
    print("開始更新…")
    mismatch = False
    for name, old, new, expected in rows:
        result = db[name].update_many(
            {"research_project_code": old},
            {"$set": {"research_project_code": new}},
        )
        flag = ""
        if result.modified_count != expected:
            flag = f"  [WARN] 與 dry-run 計數 {expected} 不符"
            mismatch = True
        print(f"{name:<10} {old:<16} → {new:<20} "
              f"matched={result.matched_count} modified={result.modified_count}{flag}")

    print("\n複掃驗證：")
    leftover = 0
    for name in COLLECTIONS:
        for old in RENAME_MAP:
            n = db[name].count_documents({"research_project_code": old})
            if n:
                print(f"  [WARN] {name} 仍有 {n} 筆 {old}")
                leftover += n
    if leftover == 0 and not mismatch:
        print("  [OK] 所有舊值已為 0 筆，更新計數與 dry-run 一致")
    elif leftover == 0:
        print("  舊值已為 0 筆，但部分計數與 dry-run 不符（見上方 WARN），請確認期間是否有其他寫入")


def main():
    parser = argparse.ArgumentParser(description="nsc96 research_project_code 改名")
    parser.add_argument("--apply", action="store_true",
                        help="實際更新 DB（預設只 dry-run）")
    args = parser.parse_args()

    db = get_db()
    rows, unexpected, new_exists = scan(db)
    print_report(rows, unexpected, new_exists)

    if not args.apply:
        write_precheck_xlsx(rows, unexpected, new_exists)
        print("\nDry-run 完成，未更動任何資料。確認無誤後加 --apply 執行更新。")
        return

    if new_exists:
        print("\n[ABORT] 新值已存在於 DB，為避免混入不同批資料，中止更新。請先釐清 warnings。")
        sys.exit(1)
    apply_rename(db, rows)


if __name__ == "__main__":
    main()
