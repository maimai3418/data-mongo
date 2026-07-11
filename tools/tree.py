import os
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def collect_tree(path, ignore, depth=0, rows=None):
    if rows is None:
        rows = []
    entries = sorted(os.listdir(path))
    entries = [e for e in entries if e not in ignore]
    for name in entries:
        full = os.path.join(path, name)
        is_dir = os.path.isdir(full)
        ext = "" if is_dir else os.path.splitext(name)[1]
        size = "" if is_dir else os.path.getsize(full)
        rows.append({
            "depth": depth,
            "name": name,
            "type": "📁 Folder" if is_dir else "📄 File",
            "ext": ext,
            "size": size,
            "path": os.path.relpath(full, rows[0]["_root"] if rows else path)
        })
        if not rows[0:1] or "_root" not in rows[0]:
            rows[-1]["_root"] = path
        else:
            rows[-1]["_root"] = rows[0]["_root"]
        if is_dir:
            collect_tree(full, ignore, depth + 1, rows)
    return rows

def build_xlsx(root, rows, output):
    wb = Workbook()
    ws = wb.active
    ws.title = "Directory Tree"

    headers = ["Name", "Type", "Extension", "Size (bytes)", "Relative Path"]
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    folder_font = Font(bold=True, name="Arial", size=10)
    file_font = Font(name="Arial", size=10)
    folder_fill = PatternFill("solid", fgColor="D9E2F3")

    for i, r in enumerate(rows, 2):
        indent = "    " * r["depth"]
        ws.cell(row=i, column=1, value=f"{indent}{r['name']}").font = folder_font if r["type"].startswith("📁") else file_font
        ws.cell(row=i, column=2, value=r["type"])
        ws.cell(row=i, column=3, value=r["ext"])
        ws.cell(row=i, column=4, value=r["size"] if r["size"] != "" else "")
        ws.cell(row=i, column=5, value=r["path"])
        if r["type"].startswith("📁"):
            for col in range(1, 6):
                ws.cell(row=i, column=col).fill = folder_fill

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 60
    ws.auto_filter.ref = f"A1:E{len(rows)+1}"
    wb.save(output)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export folder structure to Excel")
    parser.add_argument("path", help="Target folder path")
    parser.add_argument("-o", "--output", default="directory_tree.xlsx", help="Output xlsx path")
    parser.add_argument("--ignore", nargs="*", default=[".git", "node_modules", "__pycache__", ".venv"])
    args = parser.parse_args()

    root = os.path.abspath(args.path)
    if not os.path.isdir(root):
        print(f"Error: {root} is not a valid directory")
        exit(1)

    rows = [{"depth": 0, "name": os.path.basename(root), "type": "📁 Folder", "ext": "", "size": "", "path": ".", "_root": root}]
    collect_tree(root, args.ignore, depth=1, rows=rows)
    for r in rows:
        r.pop("_root", None)

    build_xlsx(root, rows, args.output)
    print(f"✅ Exported {len(rows)} items → {args.output}")