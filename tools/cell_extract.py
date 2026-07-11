#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
掃描指定資料夾下所有 xls 檔案，依「儲存格位址 → 輸出欄名」對應表，
讀出每個檔案的特定儲存格數值，整理成一份新的 xlsx（每個檔案一列）。

對應表由使用者自行定義：以 Excel 儲存格位址（A1、B5…）指定要讀的格子，
並指定它在輸出檔案中的欄位名稱。

用法：
  # 直接在指令列定義對應（CELL=欄名，可給多組）
  python tools/cell_extract.py <資料夾> --map A1=id B5=name C2=score

  # 用 JSON 檔定義對應（{"A1": "id", "B5": "name"}）
  python tools/cell_extract.py <資料夾> --map-file mapping.json

  # 其他選項
  python tools/cell_extract.py <資料夾> --map A1=id -o out.xlsx
  python tools/cell_extract.py <資料夾> --map A1=id --sheet 名稱或索引
  python tools/cell_extract.py <資料夾> --map A1=id --recursive
  python tools/cell_extract.py <資料夾> --map A1=id --ext .xls .xlsx

需求套件：pandas、openpyxl、xlrd（讀 .xls 用）
"""

import os
import sys
import json
import argparse

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple

# 讓中文/emoji 在 Windows 主控台（cp950）也能正常輸出
try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

# ── CONFIG ──────────────────────────────────────────────────────────────
DEFAULT_OUTPUT = "cell_extract.xlsx"
DEFAULT_EXT = [".xls", ".xlsx", ".xlsm"]
SOURCE_COL = "source_file"   # 輸出檔的第一欄：來源檔名
NOTE_COL = "_note"           # 僅在有檔案讀取失敗時才出現
# ────────────────────────────────────────────────────────────────────────

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
BASE_FONT = Font(name=FONT_NAME)
ERR_FILL = PatternFill("solid", fgColor="FCE4D6")


def norm(v):
    """把儲存格值正規化：NaN→None、整數型 float→int、字串去頭尾空白。"""
    try:
        if v is None or pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str):
        return v.strip()
    return v


def parse_inline_map(pairs):
    """把 ['A1=id', 'B5=name'] 解析成 [(cell, name), ...]，保留順序。"""
    mapping = []
    for token in pairs or []:
        if "=" not in token:
            raise SystemExit(f"對應格式錯誤：「{token}」，應為 CELL=欄名（例：A1=id）")
        cell, name = token.split("=", 1)
        cell, name = cell.strip(), name.strip()
        if not cell or not name:
            raise SystemExit(f"對應格式錯誤：「{token}」，CELL 與欄名都不可為空")
        mapping.append((cell, name))
    return mapping


def load_map_file(path):
    """讀取 JSON 對應檔（{"A1": "id", "B5": "name"}），保留順序。"""
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, dict):
        raise SystemExit(f"對應檔 {path} 內容須為 JSON 物件，例：{{\"A1\": \"id\"}}")
    return [(str(cell).strip(), str(name).strip()) for cell, name in data.items()]


def build_mapping(args):
    """合併 --map-file 與 --map（inline 在後，可覆寫同位址），驗證並回傳有序清單。"""
    mapping = []
    if args.map_file:
        mapping.extend(load_map_file(args.map_file))
    mapping.extend(parse_inline_map(args.map))

    # 驗證每個儲存格位址；同位址後者覆寫前者但保留首次出現順序
    seen = {}
    ordered = []
    for cell, name in mapping:
        try:
            coordinate_to_tuple(cell.upper())
        except (ValueError, TypeError):
            raise SystemExit(f"儲存格位址無效：「{cell}」（例：A1、B5）")
        if cell.upper() in seen:
            ordered[seen[cell.upper()]] = (cell.upper(), name)
        else:
            seen[cell.upper()] = len(ordered)
            ordered.append((cell.upper(), name))

    if not ordered:
        raise SystemExit("未提供任何對應，請用 --map A1=id 或 --map-file 指定。")

    # 輸出欄名不可重複
    names = [name for _, name in ordered]
    dup = {n for n in names if names.count(n) > 1}
    if dup:
        raise SystemExit(f"輸出欄名重複：{sorted(dup)}，請改用不同名稱。")
    return ordered


def resolve_sheet(arg):
    """--sheet 為純數字→當索引（int）；否則當工作表名稱；省略→第一個工作表（0）。"""
    if arg is None:
        return 0
    return int(arg) if arg.lstrip("-").isdigit() else arg


def collect_files(root, exts, recursive):
    """蒐集資料夾下符合副檔名的檔案（排序固定，避免每次順序不同）。"""
    exts = {e.lower() for e in exts}
    found = []
    if recursive:
        for dirpath, _, filenames in os.walk(root):
            for fname in filenames:
                if os.path.splitext(fname)[1].lower() in exts:
                    found.append(os.path.join(dirpath, fname))
    else:
        for fname in os.listdir(root):
            fpath = os.path.join(root, fname)
            if os.path.isfile(fpath) and os.path.splitext(fname)[1].lower() in exts:
                found.append(fpath)
    return sorted(found)


def get_cell(df, cell):
    """依儲存格位址從 header=None 的 DataFrame 取值，超出範圍回傳 None。"""
    row, col = coordinate_to_tuple(cell)   # 皆為 1-based
    if row - 1 >= len(df) or col - 1 >= df.shape[1]:
        return None
    return norm(df.iat[row - 1, col - 1])


def extract_file(path, sheet, mapping):
    """讀單一檔案指定工作表，回傳 {欄名: 值} 與錯誤訊息（成功時為 None）。"""
    try:
        df = pd.read_excel(path, sheet_name=sheet, header=None, dtype=object)
    except Exception as e:                 # 讀檔/工作表失敗：保留該列，記錄錯誤
        return {name: None for _, name in mapping}, str(e)
    return {name: get_cell(df, cell) for cell, name in mapping}, None


def write_xlsx(rows, out_names, output, has_error):
    """把每列資料寫成有樣式的 xlsx。"""
    headers = [SOURCE_COL] + out_names + ([NOTE_COL] if has_error else [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted"

    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, row in enumerate(rows, start=2):
        is_err = bool(row.get(NOTE_COL))
        for j, h in enumerate(headers, start=1):
            cell = ws.cell(row=i, column=j, value=row.get(h))
            cell.font = BASE_FONT
            if is_err:
                cell.fill = ERR_FILL

    # 欄寬自動調整
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        ws.column_dimensions[letter].width = min(max(longest + 2, 12), 60)

    ws.freeze_panes = "A2"
    if len(rows) >= 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    wb.save(output)


def main():
    parser = argparse.ArgumentParser(
        description="讀取資料夾內所有 xls 的特定儲存格，整理輸出成新的 xlsx")
    parser.add_argument("folder", help="要掃描的資料夾路徑")
    parser.add_argument("--map", nargs="+", metavar="CELL=NAME", default=None,
                        help="儲存格位址→欄名對應，可多組（例：--map A1=id B5=name）")
    parser.add_argument("--map-file", metavar="JSON",
                        help='JSON 對應檔，內容如 {"A1": "id", "B5": "name"}')
    parser.add_argument("--sheet", default=None,
                        help="工作表名稱或索引（純數字視為索引；省略=第一個工作表）")
    parser.add_argument("-r", "--recursive", action="store_true",
                        help="遞迴掃描子資料夾（預設只掃最外層）")
    parser.add_argument("--ext", nargs="+", default=DEFAULT_EXT, metavar="EXT",
                        help=f"要讀取的副檔名（預設 {' '.join(DEFAULT_EXT)}）")
    parser.add_argument("-o", "--output", default=DEFAULT_OUTPUT,
                        help=f"輸出檔名（預設 {DEFAULT_OUTPUT}）")
    args = parser.parse_args()

    root = os.path.abspath(args.folder)
    if not os.path.isdir(root):
        print(f"Error: 找不到資料夾 {root}")
        sys.exit(1)

    mapping = build_mapping(args)
    out_names = [name for _, name in mapping]
    sheet = resolve_sheet(args.sheet)
    exts = [e if e.startswith(".") else "." + e for e in args.ext]

    files = collect_files(root, exts, args.recursive)
    if not files:
        print(f"Error: {root} 內找不到副檔名為 {exts} 的檔案")
        sys.exit(1)

    print(f"🔍 掃描：{root}（{'遞迴' if args.recursive else '僅最外層'}）")
    print(f"📄 檔案數：{len(files)}｜工作表：{sheet}")
    print(f"🗺️  對應：{', '.join(f'{c}->{n}' for c, n in mapping)}")

    rows, n_error = [], 0
    for path in files:
        values, err = extract_file(path, sheet, mapping)
        row = {SOURCE_COL: os.path.relpath(path, root)}
        row.update(values)
        if err:
            n_error += 1
            row[NOTE_COL] = f"[讀取失敗] {err}"
            print(f"   ⚠️  {row[SOURCE_COL]}：{err}")
        rows.append(row)

    write_xlsx(rows, out_names, args.output, has_error=n_error > 0)

    print(f"✅ 完成：{len(rows)} 列 → {os.path.abspath(args.output)}")
    print(f"   欄位（{len(out_names) + 1}）：{[SOURCE_COL] + out_names}")
    if n_error:
        print(f"   ⚠️  有 {n_error} 個檔案讀取失敗（已於 {NOTE_COL} 欄標註並標色）")


if __name__ == "__main__":
    main()
