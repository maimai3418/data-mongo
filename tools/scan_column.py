import os
import sys
import argparse
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 讓本腳本可從 tools/ 子目錄被直接執行：把專案根目錄加入 sys.path，
# 以便 import 根目錄的 src 套件（wait_and_retry 等）。
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from src.utils.wait_and_retry import wait_and_retry

SUPPORTED_EXT = {".xlsx", ".xls", ".xlsm", ".csv", ".tsv"}
KEYWORDS = ["time", "test", "study"]

def get_columns(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    try:
        if ext == ".csv":
            df = pd.read_csv(filepath, nrows=0, encoding="utf-8", on_bad_lines="skip")
            return {"Sheet1": list(df.columns)}
        elif ext == ".tsv":
            df = pd.read_csv(filepath, nrows=0, sep="\t", encoding="utf-8", on_bad_lines="skip")
            return {"Sheet1": list(df.columns)}
        else:
            sheets = pd.read_excel(filepath, sheet_name=None, nrows=0)
            return {name: list(df.columns) for name, df in sheets.items()}
    except Exception as e:
        return {"_error": [str(e)]}

def match_keywords(columns, keywords):
    matched = []
    for col in columns:
        col_lower = str(col).lower()
        if any(kw in col_lower for kw in keywords):
            matched.append(str(col))
    return matched

def scan_folder(root, ignore, keywords):
    results = []
    for dirpath, dirnames, filenames in os.walk(root):
        dirnames[:] = [d for d in dirnames if d not in ignore]
        for fname in filenames:
            ext = os.path.splitext(fname)[1].lower()
            if ext not in SUPPORTED_EXT:
                continue
            fpath = os.path.join(dirpath, fname)
            rel = os.path.relpath(fpath, root)
            sheets = get_columns(fpath)
            if "_error" in sheets:
                results.append({"file": fname, "path": rel, "sheet": "-", "matched": f"[Error] {sheets['_error'][0]}", "all_cols": ""})
                continue
            for sheet_name, cols in sheets.items():
                matched = match_keywords(cols, keywords)
                if matched:
                    results.append({
                        "file": fname,
                        "path": rel,
                        "sheet": sheet_name,
                        "matched": ", ".join(matched),
                        "all_cols": ", ".join(str(c) for c in cols)
                    })
    return results

def write_xlsx(results, output):
    wb = Workbook()
    ws = wb.active
    ws.title = "Matched Files"

    headers = ["File Name", "Relative Path", "Sheet", "Matched Columns", "All Columns"]
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    hfill = PatternFill("solid", fgColor="4472C4")
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill, c.alignment = hfont, hfill, Alignment(horizontal="center")

    match_fill = PatternFill("solid", fgColor="FFF2CC")
    base_font = Font(name="Arial", size=10)
    for r, row in enumerate(results, 2):
        ws.cell(row=r, column=1, value=row["file"]).font = base_font
        ws.cell(row=r, column=2, value=row["path"]).font = base_font
        ws.cell(row=r, column=3, value=row["sheet"]).font = base_font
        mc = ws.cell(row=r, column=4, value=row["matched"])
        mc.font = Font(bold=True, name="Arial", size=10, color="C00000")
        mc.fill = match_fill
        ws.cell(row=r, column=5, value=row["all_cols"]).font = base_font

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 60
    ws.auto_filter.ref = f"A1:E{len(results)+1}"
    wait_and_retry(lambda: wb.save(output), output)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Scan folder for files with specific column keywords")
    parser.add_argument("path", help="Root folder to scan")
    parser.add_argument("-o", "--output", default="matched_files.xlsx")
    parser.add_argument("--keywords", nargs="*", default=KEYWORDS, help="Column name keywords to match (default: time test study)")
    parser.add_argument("--ignore", nargs="*", default=[".git", "node_modules", "__pycache__", ".venv"])
    args = parser.parse_args()

    root = os.path.abspath(args.path)
    if not os.path.isdir(root):
        print(f"Error: {root} is not a valid directory")
        sys.exit(1)

    kw = [k.lower() for k in args.keywords]
    print(f"🔍 Scanning: {root}")
    print(f"🔑 Keywords: {kw}")
    results = scan_folder(root, args.ignore, kw)
    write_xlsx(results, args.output)
    print(f"✅ Found {len(results)} matches → {args.output}")