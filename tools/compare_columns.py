#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
比對多個 xlsx/csv 檔案（所有工作表）的標題列（head row）。

讀取指定檔案或資料夾內所有 xlsx/xls/csv 的每個工作表，只取標題列，
輸出一份對照報表，標示每個欄位出現在哪些檔案/工作表：
  - Columns  欄位 × 來源矩陣（儲存格顯示該來源的原始欄名，
             非所有來源都有的欄位整列標黃）
  - Sources  來源清單（每個檔案｜工作表的欄位數與完整標題列，
             含讀取失敗的檔案）

欄位名稱先 strip() 轉小寫再比對（famid / famID 視為相同），
矩陣中顯示各來源的原始寫法。

用法（從專案根目錄執行，-o 輸出檔名必填）：
  python tools/compare_columns.py <檔案或資料夾>... -o 報表名稱.xlsx
  python tools/compare_columns.py data_dir/ -o columns_report.xlsx
  python tools/compare_columns.py a.xlsx b.csv -o 對照.xlsx

需求套件：pandas、openpyxl
"""

import os
import sys
import argparse
from collections import OrderedDict

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.utils.wait_and_retry import wait_and_retry

# 讓中文/emoji 在 Windows 主控台（cp950）也能正常輸出
try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

# ---- 報表樣式（沿用 cpt_precheck.py 的風格）----
FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="4472C4")
WARN_FILL = PatternFill("solid", start_color="FFF2CC")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=12)
BASE_FONT = Font(name=FONT_NAME)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WARN_MARK = "⚠️ 缺漏"


# ── 報表工具（沿用 cpt_precheck.py）─────────────────────────────────────
def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autofit(ws, max_width=50):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 0
        for cell in col:
            if cell.value is not None:
                longest = max(longest, max(len(s) for s in str(cell.value).split("\n")))
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), max_width)


def write_table(ws, headers, rows, start_row=1, warn_col=None, warn_value=None):
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h).font = BASE_FONT
    style_header(ws, len(headers), row=start_row)
    for i, rowvals in enumerate(rows, start=start_row + 1):
        flag = warn_col is not None and rowvals[warn_col] == warn_value
        for j, v in enumerate(rowvals, start=1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = BASE_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if flag:
                cell.fill = WARN_FILL
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


# ── 檔案蒐集（沿用 cpt_precheck.py）─────────────────────────────────────
def collect_files(paths, output_path):
    files = []
    out_ap = os.path.abspath(output_path)
    for p in paths:
        if os.path.isdir(p):
            for f in sorted(os.listdir(p)):
                if f.lower().endswith((".xlsx", ".xls", ".csv")) and not f.startswith("~$"):
                    files.append(os.path.join(p, f))
        elif os.path.isfile(p):
            files.append(p)
        else:
            print(f"⚠️  略過：找不到 {p}")
    seen, unique = set(), []
    for f in files:
        ap = os.path.abspath(f)
        if ap in seen or ap == out_ap:
            continue
        seen.add(ap)
        unique.append(f)
    return unique


# ── 讀取標題列 ──────────────────────────────────────────────────────────
CSV_ENCODINGS = ("utf-8-sig", "cp950")  # 依序嘗試（BOM utf-8 → Windows 繁中）


def read_csv_header(filepath):
    """讀取 csv 標題列，依 CSV_ENCODINGS 依序嘗試編碼。"""
    last_err = None
    for enc in CSV_ENCODINGS:
        try:
            return pd.read_csv(filepath, nrows=0, encoding=enc)
        except UnicodeDecodeError as e:
            last_err = e
    raise last_err


def read_headers(files):
    """讀取所有檔案所有工作表的標題列（nrows=0 只讀標題、不載入資料）。

    回傳 (sources, read_errors)：
      sources     OrderedDict {來源(檔名[｜工作表]): [原始欄名, ...]}
      read_errors [(檔名, 錯誤訊息)]
    """
    sources = OrderedDict()
    read_errors = []
    for filepath in files:
        filename = os.path.basename(filepath)
        try:
            if filename.lower().endswith(".csv"):
                all_sheets = {None: read_csv_header(filepath)}
            else:
                all_sheets = pd.read_excel(filepath, sheet_name=None, nrows=0)
        except Exception as e:
            read_errors.append((filename, f"無法讀取檔案: {e}"))
            continue
        multi = len(all_sheets) > 1
        for sheet_name, df in all_sheets.items():
            source = f"{filename}｜{sheet_name}" if multi else filename
            sources[source] = [str(c) for c in df.columns]
    return sources, read_errors


def build_matrix(sources):
    """彙整所有來源的欄位。

    回傳 (欄位順序清單(小寫)、{小寫欄名: {來源: 原始欄名}})。
    """
    order = []
    cells = {}
    for source, headers in sources.items():
        for h in headers:
            n = h.strip().lower()
            if n not in cells:
                cells[n] = OrderedDict()
                order.append(n)
            # 同一來源出現多個同名欄位時全部列出
            if source in cells[n]:
                cells[n][source] += f", {h}"
            else:
                cells[n][source] = h
    return order, cells


# ── 輸出 ────────────────────────────────────────────────────────────────
def build_report(output_path, files, sources, read_errors):
    order, cells = build_matrix(sources)
    src_list = list(sources.keys())
    n_missing = sum(1 for n in order if len(cells[n]) < len(src_list))

    wb = Workbook()

    # ---- Columns（欄位 × 來源矩陣）----
    ws = wb.active
    ws.title = "Columns"
    ws["A1"] = "標題列（head row）對照"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (f"來源 {len(src_list)} 個（{len(files)} 個檔案）｜"
                f"欄位聯集 {len(order)} 個｜欄名以 strip+小寫比對，"
                f"儲存格顯示各來源的原始欄名")
    ws["A2"].font = BASE_FONT
    headers = ["欄位名稱", "出現來源數", "標記", *src_list]
    rows = []
    for n in order:
        hit = cells[n]
        mark = "" if len(hit) == len(src_list) else WARN_MARK
        rows.append((n, len(hit), mark, *[hit.get(s, "") for s in src_list]))
    if not rows:
        rows = [("（無）", "", "", *[""] * len(src_list))]
    write_table(ws, headers, rows, start_row=4, warn_col=2, warn_value=WARN_MARK)
    ws.freeze_panes = "B5"  # 捲動時保留欄位名稱那一欄
    autofit(ws, max_width=40)

    # ---- Sources（來源清單）----
    ws = wb.create_sheet("Sources")
    src_rows = [(s, len(h), ", ".join(h)) for s, h in sources.items()]
    for filename, err in read_errors:
        src_rows.append((filename, 0, f"讀取失敗：{err}"))
    if not src_rows:
        src_rows = [("（無）", 0, "")]
    write_table(ws, ["檔案（｜工作表）", "欄位數", "標題列內容（原始）"], src_rows)
    autofit(ws, max_width=80)

    wait_and_retry(lambda: wb.save(output_path), output_path)
    return len(order), n_missing


def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="比對多個 xlsx/csv 檔案（所有工作表）的標題列，輸出欄位×來源對照表")
    parser.add_argument("paths", nargs="+",
                        help="要讀取的 xlsx/csv 檔案或資料夾")
    parser.add_argument("-o", "--output", required=True, metavar="XLSX",
                        help="報表輸出檔名（必填，例如 columns_report.xlsx）")
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)

    files = collect_files(args.paths, args.output)
    if not files:
        print("Error: 找不到任何 xlsx/csv 檔案")
        sys.exit(1)
    print(f"📄 檔案數：{len(files)}")

    sources, read_errors = read_headers(files)
    for filename, err in read_errors:
        print(f"⚠️  {filename}: {err}")

    n_cols, n_missing = build_report(args.output, files, sources, read_errors)

    print("\n===== 對照結果 =====")
    print(f"來源（檔案｜工作表）數：{len(sources)}")
    print(f"讀取失敗檔案數：{len(read_errors)}")
    print(f"欄位聯集總數：{n_cols}")
    print(f"非所有來源都有的欄位數：{n_missing}")
    print(f"\n已輸出報表：{args.output}")


if __name__ == "__main__":
    main()
