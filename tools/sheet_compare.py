#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
比對「同一個 xlsx 檔案」內兩個工作表（A、B）的欄位與各 row 數值異同。

對齊方式（決定 A、B 的某一列要跟對方哪一列比）：
  1) 指定 --key 欄位：以該欄位的值當作鍵對齊（建議，順序不同也能比）。
     可指定多個欄位組成複合鍵，例如 --key famid role。
  2) 不指定 --key：以「列的位置」對齊（A 第 1 筆 對 B 第 1 筆 ...）。

輸出：comparison.xlsx，含 5 個工作表
  1) Summary       彙總
  2) Columns       欄位比較（共同 / 只在 A / 只在 B）
  3) Cell_Diff     逐列逐欄的值差異（僅比共同欄位、可對齊的列）
  4) Only_in_A     只在 A 出現的整筆資料
  5) Only_in_B     只在 B 出現的整筆資料

範例：
  python tools/sheet_compare.py data.xlsx                      # 預設取前兩個工作表，依位置對齊
  python tools/sheet_compare.py data.xlsx --sheets A B         # 指定工作表名稱
  python tools/sheet_compare.py data.xlsx --key famid          # 依 famid 欄位對齊
  python tools/sheet_compare.py data.xlsx --key famid role      # 依複合鍵對齊
  python tools/sheet_compare.py data.xlsx -o out.xlsx

需求套件：pandas、openpyxl
"""

import os
import sys
import argparse
from collections import Counter

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 讓中文/emoji 在 Windows 主控台（cp1252/cp950）也能正常輸出
try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="4472C4")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=12)
BASE_FONT = Font(name=FONT_NAME)
DIFF_FILL = PatternFill("solid", start_color="FFF2CC")
ONLY_A_FILL = PatternFill("solid", start_color="DDEBF7")
ONLY_B_FILL = PatternFill("solid", start_color="FCE4D6")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

EMPTY = "<空>"          # 顯示用：空值 / NaN
MISSING = "<無此欄位>"  # 顯示用：該表沒有這個欄位


# ---------------------------------------------------------------- 讀取 / 正規化

def norm(v):
    """把儲存格值正規化成可比較的形式（NaN→None、整數型 float→int、字串去空白）。"""
    try:
        if v is None or pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str):
        return v.strip()
    return v


def disp(v):
    """顯示用：None / 空字串→<空>。"""
    n = norm(v)
    if n is None or n == "":
        return EMPTY
    return n


def load_sheets(path, sheets):
    """讀取兩個指定工作表，回傳 (name_a, df_a, name_b, df_b)。"""
    xls = pd.ExcelFile(path)
    avail = xls.sheet_names
    if sheets:
        for s in sheets:
            if s not in avail:
                raise SystemExit(
                    f"找不到工作表「{s}」。檔案內現有工作表：{avail}")
        name_a, name_b = sheets
    else:
        if len(avail) < 2:
            raise SystemExit(f"檔案只有 {len(avail)} 個工作表，無法比對：{avail}")
        name_a, name_b = avail[0], avail[1]
    df_a = pd.read_excel(xls, sheet_name=name_a, dtype=object)
    df_b = pd.read_excel(xls, sheet_name=name_b, dtype=object)
    return name_a, df_a, name_b, df_b


# ---------------------------------------------------------------- 欄位比較

def compare_columns(cols_a, cols_b):
    set_a, set_b = set(cols_a), set(cols_b)
    common = [c for c in cols_a if c in set_b]
    only_a = [c for c in cols_a if c not in set_b]
    only_b = [c for c in cols_b if c not in set_a]
    return common, only_a, only_b


# ---------------------------------------------------------------- 列對齊

def build_key_index(df, keys):
    """以 keys（一或多個欄位）的值組成複合鍵對齊，回傳 {key_tuple: row_dict} 與重複鍵清單。"""
    index, dups = {}, Counter()
    for _, row in df.iterrows():
        k = tuple(norm(row.get(c)) for c in keys)
        dups[k] += 1
        if k not in index:
            index[k] = row.to_dict()
    return index, {k: n for k, n in dups.items() if n > 1}


def fmt_key(keys, key_tuple):
    """把複合鍵格式化成可讀字串：單欄直接顯示值，多欄顯示「欄位=值 | 欄位=值」。"""
    if len(keys) == 1:
        return disp(key_tuple[0])
    return " | ".join(f"{c}={disp(v)}" for c, v in zip(keys, key_tuple))


def diff_by_key(df_a, df_b, keys, common_cols):
    idx_a, dups_a = build_key_index(df_a, keys)
    idx_b, dups_b = build_key_index(df_b, keys)
    keys_a, keys_b = set(idx_a), set(idx_b)
    both = keys_a & keys_b
    keyset = set(keys)
    cmp_cols = [c for c in common_cols if c not in keyset]

    cell_diffs = []
    for k in sorted(both, key=lambda t: tuple(str(x) for x in t)):
        ra, rb = idx_a[k], idx_b[k]
        for c in cmp_cols:
            va, vb = norm(ra.get(c)), norm(rb.get(c))
            if va != vb:
                cell_diffs.append((fmt_key(keys, k), c, disp(ra.get(c)), disp(rb.get(c))))

    only_a = [idx_a[k] for k in idx_a if k not in keys_b]
    only_b = [idx_b[k] for k in idx_b if k not in keys_a]
    return cell_diffs, only_a, only_b, dups_a, dups_b, len(both)


def diff_by_position(df_a, df_b, common_cols):
    n = min(len(df_a), len(df_b))
    cell_diffs = []
    for i in range(n):
        ra, rb = df_a.iloc[i], df_b.iloc[i]
        for c in common_cols:
            va, vb = norm(ra.get(c)), norm(rb.get(c))
            if va != vb:
                # 列號以 Excel 資料列表示（標題列為第 1 列，故 +2）
                cell_diffs.append((f"列 {i + 2}", c, disp(ra.get(c)), disp(rb.get(c))))
    only_a = [df_a.iloc[i].to_dict() for i in range(n, len(df_a))]
    only_b = [df_b.iloc[i].to_dict() for i in range(n, len(df_b))]
    return cell_diffs, only_a, only_b, n


# ---------------------------------------------------------------- 寫出 xlsx

def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autofit(ws, max_width=60):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 0
        for cell in col:
            if cell.value is not None:
                longest = max(longest, max(len(s) for s in str(cell.value).split("\n")))
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), max_width)


def write_table(ws, headers, rows, start_row=1, row_fill=None):
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h).font = BASE_FONT
    style_header(ws, len(headers), row=start_row)
    for i, rowvals in enumerate(rows, start=start_row + 1):
        for j, v in enumerate(rowvals, start=1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = BASE_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_fill is not None:
                cell.fill = row_fill
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


def write_records(ws, records, cols):
    """把 row_dict 清單依 cols 順序寫成表格。"""
    if not records:
        write_table(ws, ["（無）"], [["沒有資料"]])
        autofit(ws)
        return
    rows = [[disp(r.get(c)) for c in cols] for r in records]
    write_table(ws, list(cols), rows)
    autofit(ws)


def main():
    parser = argparse.ArgumentParser(
        description="比對同一個 xlsx 檔內兩個工作表的欄位與各 row 數值異同")
    parser.add_argument("file", help="要比對的 xlsx 檔案路徑")
    parser.add_argument("--sheets", nargs=2, metavar=("A", "B"),
                        help="兩個工作表名稱（預設取檔案中的前兩個）")
    parser.add_argument("--key", nargs="+", default=None, metavar="COL",
                        help="對齊用的鍵欄位，可指定多個組成複合鍵"
                             "（例：--key famid role；省略則以列的位置對齊）")
    parser.add_argument("-o", "--output", default="comparison.xlsx",
                        help="輸出檔名（預設 comparison.xlsx）")
    args = parser.parse_args()

    path = os.path.abspath(args.file)
    if not os.path.isfile(path):
        print(f"Error: 找不到檔案 {path}")
        sys.exit(1)

    name_a, df_a, name_b, df_b = load_sheets(path, args.sheets)
    cols_a, cols_b = list(df_a.columns), list(df_b.columns)
    common, only_col_a, only_col_b = compare_columns(cols_a, cols_b)

    if args.key:
        missing = [c for c in args.key if c not in cols_a or c not in cols_b]
        if missing:
            raise SystemExit(
                f"鍵欄位 {missing} 必須同時存在於兩個工作表。"
                f"\n  {name_a} 欄位：{cols_a}\n  {name_b} 欄位：{cols_b}")
        cell_diffs, only_a, only_b, dups_a, dups_b, both_n = diff_by_key(
            df_a, df_b, args.key, common)
        align_desc = f"依鍵欄位「{' | '.join(args.key)}」對齊"
        aligned_n = both_n
    else:
        cell_diffs, only_a, only_b, aligned_n = diff_by_position(df_a, df_b, common)
        dups_a, dups_b = {}, {}
        align_desc = "依列的位置對齊"

    # ---------------- 輸出
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"工作表比對彙總（A={name_a} / B={name_b}）"
    ws["A1"].font = TITLE_FONT
    summary = [
        ("項目", "數值"),
        ("檔案", os.path.basename(path)),
        ("對齊方式", align_desc),
        (f"A「{name_a}」資料筆數", len(df_a)),
        (f"B「{name_b}」資料筆數", len(df_b)),
        ("A 欄位數", len(cols_a)),
        ("B 欄位數", len(cols_b)),
        ("共同欄位數", len(common)),
        ("只在 A 的欄位數", len(only_col_a)),
        ("只在 B 的欄位數", len(only_col_b)),
        ("可對齊比對的列數", aligned_n),
        ("值有差異的儲存格數", len(cell_diffs)),
        ("值有差異的列數", len({d[0] for d in cell_diffs})),
        ("只在 A 的列數", len(only_a)),
        ("只在 B 的列數", len(only_b)),
    ]
    if args.key:
        summary.append(("A 重複鍵種類數", len(dups_a)))
        summary.append(("B 重複鍵種類數", len(dups_b)))
    write_table(ws, summary[0], summary[1:], start_row=3)
    autofit(ws)

    ws = wb.create_sheet("Columns")
    col_rows = []
    for c in common:
        col_rows.append((c, "✔", "✔", "共同"))
    for c in only_col_a:
        col_rows.append((c, "✔", "", f"只在 A（{name_a}）"))
    for c in only_col_b:
        col_rows.append((c, "", "✔", f"只在 B（{name_b}）"))
    write_table(ws, ["欄位", f"A：{name_a}", f"B：{name_b}", "狀態"], col_rows)
    autofit(ws)

    ws = wb.create_sheet("Cell_Diff")
    key_label = " | ".join(args.key) if args.key else "列位置"
    if cell_diffs:
        write_table(ws,
                    [key_label, "欄位", f"A 值（{name_a}）", f"B 值（{name_b}）"],
                    cell_diffs, row_fill=DIFF_FILL)
    else:
        write_table(ws, [key_label, "欄位", f"A 值（{name_a}）", f"B 值（{name_b}）"],
                    [("（無）", "共同欄位的可對齊列數值皆相同", "", "")])
    autofit(ws)

    ws = wb.create_sheet("Only_in_A")
    write_records(ws, only_a, cols_a)
    ws = wb.create_sheet("Only_in_B")
    write_records(ws, only_b, cols_b)

    wb.save(args.output)

    print(f"✅ 已輸出：{os.path.abspath(args.output)}")
    print(f"   A={name_a}（{len(df_a)} 列）｜B={name_b}（{len(df_b)} 列）｜{align_desc}")
    print(f"   共同欄位 {len(common)}｜只在A欄位 {len(only_col_a)}｜只在B欄位 {len(only_col_b)}")
    print(f"   差異儲存格 {len(cell_diffs)}（{len({d[0] for d in cell_diffs})} 列）"
          f"｜只在A列 {len(only_a)}｜只在B列 {len(only_b)}")


if __name__ == "__main__":
    main()
